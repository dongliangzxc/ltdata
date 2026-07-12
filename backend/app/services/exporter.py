"""
导出服务：基于型号匹配结果，按品类分 Sheet 导出。
- 已匹配/已确认的条目 → 按 category_name 分 Sheet，含动态规格列
- 待确认条目 → 单独"待确认" Sheet，无规格列
- URL 映射待确认（text_only）→ 按品类分 Sheet，命名"{品类}-URL映射待确认"
- 争议复核 / 已排除 → 分别一个总 Sheet，基础列
- 干扰项过滤 → 单独 Sheet，基础列 + 命中关键词/规则/原因；仅导出未恢复项
- 规格列按品类过滤：每个 Sheet 只显示本品类（category_code）的规格列
- 约定：models.category_name 与 metadata_specs.category_code 使用相同的品类码（如 SOUNDBAR）
- 规格值从 model_specs 查询
"""
import uuid
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.models.schemas import (
    MatchResult, RawDataRecord, ModelRecord,
    ModelSpec, MetadataSpec,
    Category, FilteredItem,
)
from app.core.config import settings

# 基础列：字段名 → 中文表头
BASE_COLS = [
    ("platform",      "平台"),
    ("month",         "月"),
    ("category_lv1",  "Lv1类目名称"),
    ("category_lv2",  "Lv2类目名称"),
    ("category_lv3",  "Lv3类目名称"),
    ("category_lv4",  "Lv4类目名称"),
    ("category_lv5",  "Lv5类目名称"),
    ("item_id",       "宝贝ID"),
    ("item_url",      "宝贝链接"),
    ("item_name",     "宝贝名称"),
    ("item_image",    "宝贝图片"),
    ("ref_price",     "参考价格"),
    ("brand_raw",     "宝贝品牌"),
    ("shop_name",     "宝贝店铺名称"),
    ("sales_qty",     "销量"),
    ("sales_amount",  "销售额"),
    ("price",         "价格"),
    ("brand_std",     "品牌"),
    ("model_code",    "型号"),
    ("brand_name",    "品牌名称"),
    ("model_name",    "型号名称"),
]

BASE_FIELD_NAMES = [f for f, _ in BASE_COLS]
BASE_CN_NAMES    = [cn for _, cn in BASE_COLS]
PAGE_SIZE = 5000


def _sheet_name(name: str, used_names: set[str]) -> str:
    base = (name or "Sheet")[:31]
    candidate = base
    index = 2
    while candidate in used_names:
        suffix = f"_{index}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used_names.add(candidate)
    return candidate


def _base_row(rd: RawDataRecord, model: ModelRecord | None = None) -> dict:
    row = {}
    for field in BASE_FIELD_NAMES:
        if field == "brand_std":
            row[field] = rd.brand_std or rd.brand_raw or ""
        elif field == "model_code":
            row[field] = model.model_code if model else ""
        elif field == "brand_name":
            row[field] = model.brand_name if model else ""
        elif field == "model_name":
            row[field] = model.model_name if model else ""
        else:
            row[field] = getattr(rd, field, None)
    return row


def _spec_map_for_models(db: Session, model_ids: list[int]) -> dict[int, dict[str, str]]:
    spec_map: dict[int, dict[str, str]] = {}
    if not model_ids:
        return spec_map
    spec_rows = db.query(ModelSpec).filter(ModelSpec.model_id.in_(model_ids)).all()
    for s in spec_rows:
        spec_map.setdefault(s.model_id, {})[s.spec_name] = s.spec_value or ""
    return spec_map


def export_match_job(
    db: Session,
    clean_job_id: int,
    filename_prefix: str = "已处理数据",
) -> list[dict]:
    """
    生成导出文件，返回 [{"filename": ..., "token": ..., "path": ..., "rows": ..., "pending_rows": ...}]
    """
    cat_map: dict[str, str] = {
        c.code: c.name
        for c in db.query(Category).all()
    }

    all_spec_defs = db.query(MetadataSpec).order_by(MetadataSpec.id).all()
    category_spec_names: dict[str, list[str]] = {}
    for s in all_spec_defs:
        category_spec_names.setdefault(s.category_code, []).append(s.spec_name)

    matched_query = (
        db.query(MatchResult, RawDataRecord, ModelRecord)
        .join(RawDataRecord, MatchResult.raw_data_id == RawDataRecord.id)
        .join(ModelRecord, MatchResult.model_id == ModelRecord.id)
        .filter(
            MatchResult.clean_job_id == clean_job_id,
            MatchResult.match_status.in_(["url_matched", "matched", "confirmed"]),
            MatchResult.is_disabled == 0,
        )
        .order_by(MatchResult.id)
    )
    pending_query = (
        db.query(MatchResult, RawDataRecord)
        .join(RawDataRecord, MatchResult.raw_data_id == RawDataRecord.id)
        .filter(
            MatchResult.clean_job_id == clean_job_id,
            MatchResult.match_status == "pending",
            MatchResult.is_disabled == 0,
        )
        .order_by(MatchResult.id)
    )
    text_only_query = (
        db.query(MatchResult, RawDataRecord, ModelRecord)
        .join(RawDataRecord, MatchResult.raw_data_id == RawDataRecord.id)
        .join(ModelRecord, MatchResult.model_id == ModelRecord.id)
        .filter(
            MatchResult.clean_job_id == clean_job_id,
            MatchResult.match_status == "text_only",
            MatchResult.is_disabled == 0,
        )
        .order_by(MatchResult.id)
    )
    # 争议复核 / 已排除：可能未绑定 model_id，使用外连接
    disputed_query = (
        db.query(MatchResult, RawDataRecord, ModelRecord)
        .join(RawDataRecord, MatchResult.raw_data_id == RawDataRecord.id)
        .outerjoin(ModelRecord, MatchResult.model_id == ModelRecord.id)
        .filter(
            MatchResult.clean_job_id == clean_job_id,
            MatchResult.match_status == "disputed",
            MatchResult.is_disabled == 0,
        )
        .order_by(MatchResult.id)
    )
    excluded_query = (
        db.query(MatchResult, RawDataRecord, ModelRecord)
        .join(RawDataRecord, MatchResult.raw_data_id == RawDataRecord.id)
        .outerjoin(ModelRecord, MatchResult.model_id == ModelRecord.id)
        .filter(
            MatchResult.clean_job_id == clean_job_id,
            MatchResult.match_status == "excluded",
            MatchResult.is_disabled == 0,
        )
        .order_by(MatchResult.id)
    )
    # 干扰项过滤：来自独立表 filtered_items，仅导出未恢复条目
    filtered_query = (
        db.query(FilteredItem, RawDataRecord)
        .join(RawDataRecord, FilteredItem.raw_data_id == RawDataRecord.id)
        .filter(
            FilteredItem.clean_job_id == clean_job_id,
            FilteredItem.is_recovered == 0,
        )
        .order_by(FilteredItem.id)
    )

    matched_total = matched_query.count()
    pending_total = pending_query.count()
    text_only_total = text_only_query.count()
    disputed_total = disputed_query.count()
    excluded_total = excluded_query.count()
    filtered_total = filtered_query.count()
    if (matched_total == 0 and pending_total == 0 and text_only_total == 0
            and disputed_total == 0 and excluded_total == 0 and filtered_total == 0):
        return []

    export_dir = Path(settings.EXPORT_DIR)
    token = uuid.uuid4().hex
    safe_name = f"{filename_prefix}.xlsx"
    file_path = export_dir / f"{token}_{safe_name}"

    used_sheet_names: set[str] = set()
    category_sheets: dict[str, tuple] = {}
    text_only_sheets: dict[str, tuple] = {}
    workbook = Workbook(write_only=True)

    def get_category_sheet(cat: str, cat_code: str):
        if cat not in category_sheets:
            spec_names = category_spec_names.get(cat_code, [])
            worksheet = workbook.create_sheet(title=_sheet_name(f"{cat}-已处理", used_sheet_names))
            worksheet.append(BASE_CN_NAMES + spec_names)
            category_sheets[cat] = (worksheet, spec_names)
        return category_sheets[cat]

    def get_text_only_sheet(cat: str):
        if cat not in text_only_sheets:
            worksheet = workbook.create_sheet(title=_sheet_name(f"{cat}-URL映射待确认", used_sheet_names))
            worksheet.append(BASE_CN_NAMES)
            text_only_sheets[cat] = (worksheet,)
        return text_only_sheets[cat][0]

    for offset in range(0, matched_total, PAGE_SIZE):
        matched_rows = matched_query.offset(offset).limit(PAGE_SIZE).all()
        spec_map = _spec_map_for_models(db, list({mr.model_id for mr, _, _ in matched_rows}))
        for mr, rd, model in matched_rows:
            row = _base_row(rd, model)
            cat_code = model.category_code or ""
            cat = cat_map.get(cat_code, cat_code) or "未知品类"
            worksheet, spec_names = get_category_sheet(cat, cat_code)
            model_specs = spec_map.get(mr.model_id, {})
            worksheet.append([row.get(field) for field in BASE_FIELD_NAMES] + [model_specs.get(sn, "") for sn in spec_names])

    pending_sheet = None
    for offset in range(0, pending_total, PAGE_SIZE):
        if pending_sheet is None:
            pending_sheet = workbook.create_sheet(title=_sheet_name("待确认", used_sheet_names))
            pending_sheet.append(BASE_CN_NAMES)
        pending_rows = pending_query.offset(offset).limit(PAGE_SIZE).all()
        for _, rd in pending_rows:
            row = _base_row(rd)
            pending_sheet.append([row.get(field) for field in BASE_FIELD_NAMES])

    for offset in range(0, text_only_total, PAGE_SIZE):
        text_only_rows = text_only_query.offset(offset).limit(PAGE_SIZE).all()
        for _, rd, model in text_only_rows:
            row = _base_row(rd, model)
            cat_code = model.category_code or ""
            cat = cat_map.get(cat_code, cat_code) or "未知品类"
            worksheet = get_text_only_sheet(cat)
            worksheet.append([row.get(field) for field in BASE_FIELD_NAMES])

    def _write_simple_match_sheet(title: str, total: int, query):
        """争议复核 / 已排除等单表 Sheet：基础列 + 可选品牌/型号（model_id 可能为空）。"""
        if total == 0:
            return
        sheet = workbook.create_sheet(title=_sheet_name(title, used_sheet_names))
        sheet.append(BASE_CN_NAMES)
        for offset in range(0, total, PAGE_SIZE):
            rows = query.offset(offset).limit(PAGE_SIZE).all()
            for _, rd, model in rows:
                row = _base_row(rd, model)
                sheet.append([row.get(field) for field in BASE_FIELD_NAMES])

    _write_simple_match_sheet("争议复核", disputed_total, disputed_query)
    _write_simple_match_sheet("已排除", excluded_total, excluded_query)

    if filtered_total > 0:
        FILTERED_EXTRA_COLS = ["命中关键词", "命中规则", "命中原因"]
        filtered_sheet = workbook.create_sheet(title=_sheet_name("干扰项过滤", used_sheet_names))
        filtered_sheet.append(BASE_CN_NAMES + FILTERED_EXTRA_COLS)
        for offset in range(0, filtered_total, PAGE_SIZE):
            filtered_rows = filtered_query.offset(offset).limit(PAGE_SIZE).all()
            for fi, rd in filtered_rows:
                row = _base_row(rd)
                filtered_sheet.append(
                    [row.get(field) for field in BASE_FIELD_NAMES]
                    + [fi.matched_keyword or "", fi.intervention_rule_name or "", fi.matched_reason or ""]
                )

    workbook.save(str(file_path))

    return [{
        "filename": safe_name,
        "token": token,
        "path": str(file_path),
        "rows": matched_total,
        "pending_rows": pending_total,
    }]
