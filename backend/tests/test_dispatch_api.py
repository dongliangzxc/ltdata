import io
import tempfile
from pathlib import Path

import openpyxl
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.api import dispatch_api
from app.api.dispatch_api import router
from app.core.auth_deps import get_current_user
from app.models.database import Base, get_db
from app.models.schemas import Category, ColumnTemplate, DispatchBatch, DispatchItem, DispatchRule, RawDataRecord, UploadFileRecord, WorkbenchExportJob


class DummyCurrentUser:
    def __init__(self, *, is_admin=1, category_permissions=None):
        self.is_admin = is_admin
        self.category_permissions = category_permissions or []


@pytest.fixture
def client_and_db(monkeypatch):
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    db = Session()
    app = FastAPI()
    app.include_router(router)

    def override_db():
        yield db

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_current_user] = lambda: DummyCurrentUser()
    monkeypatch.setattr(dispatch_api, "DISPATCH_PAGE_SIZE", 2)
    monkeypatch.setattr(dispatch_api, "DISPATCH_EXPORT_DIR", Path(tempfile.mkdtemp()))
    monkeypatch.setattr(dispatch_api, "SessionLocal", Session)
    try:
        yield TestClient(app), db
    finally:
        db.close()
        Base.metadata.drop_all(engine)


def _set_dispatch_user(client, *, is_admin=0, category_permissions=None):
    client.app.dependency_overrides[get_current_user] = lambda: DummyCurrentUser(
        is_admin=is_admin,
        category_permissions=category_permissions or [],
    )


def test_run_dispatch_processes_raw_data_in_pages(client_and_db):
    client, db = client_and_db
    file_record = UploadFileRecord(filename="large.xlsx", platform="JD", row_count=5, status="done")
    db.add(file_record)
    db.flush()
    db.add(DispatchRule(
        category_code="headphone",
        platform="jd",
        field="category_lv1",
        match_type="contains",
        value="耳机",
        priority=1,
        is_active=1,
    ))
    db.add(DispatchRule(
        category_code="speaker",
        platform=None,
        field="item_name",
        match_type="contains",
        value="音箱",
        priority=2,
        is_active=1,
    ))
    for idx, (category, item_name) in enumerate([
        ("蓝牙耳机", "商品 1"),
        ("头戴耳机", "商品 2"),
        ("智能硬件", "无线音箱"),
        ("手机配件", "保护壳"),
        ("入耳耳机", "商品 5"),
    ], start=1):
        db.add(RawDataRecord(
            file_id=file_record.id,
            platform="jd",
            month=202605,
            item_id=f"item-{idx}",
            category_lv1=category,
            item_name=item_name,
        ))
    db.commit()

    response = client.post("/api/dispatch/run", json={"file_id": file_record.id})

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "done"
    assert payload["total_rows"] == 5
    assert payload["dispatched_rows"] == 4
    assert payload["unmatched_rows"] == 1
    batch = db.query(DispatchBatch).filter_by(id=payload["id"]).one()
    assert batch.status == "done"
    items = db.query(DispatchItem).filter_by(batch_id=batch.id).all()
    assert len(items) == 4
    assert [item.category_code for item in items] == ["headphone", "headphone", "speaker", "headphone"]


def test_run_dispatch_with_category_scope_keeps_full_batch_counts(client_and_db):
    client, db = client_and_db
    file_record = UploadFileRecord(filename="scoped.xlsx", platform="JD", row_count=4, status="done")
    db.add(file_record)
    db.flush()
    db.add(DispatchRule(
        category_code="headphone",
        platform="jd",
        field="item_name",
        match_type="contains",
        value="耳机",
        priority=1,
        is_active=1,
    ))
    db.add(DispatchRule(
        category_code="speaker",
        platform="jd",
        field="item_name",
        match_type="contains",
        value="音箱",
        priority=1,
        is_active=1,
    ))
    db.flush()
    rows = [
        RawDataRecord(file_id=file_record.id, platform="JD", item_name="降噪耳机", category_lv1="耳机"),
        RawDataRecord(file_id=file_record.id, platform="JD", item_name="耳机收纳包", category_lv1="耳机"),
        RawDataRecord(file_id=file_record.id, platform="JD", item_name="蓝牙音箱", category_lv1="音箱"),
        RawDataRecord(file_id=file_record.id, platform="JD", item_name="智能门锁", category_lv1="门锁"),
    ]
    db.add_all(rows)
    db.commit()

    response = client.post("/api/dispatch/run", json={"file_id": file_record.id, "category_code": "headphone"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["total_rows"] == 4
    assert payload["dispatched_rows"] == 2
    assert payload["unmatched_rows"] == 2
    items = db.query(DispatchItem).all()
    assert [item.category_code for item in items] == ["headphone", "headphone"]


def test_run_dispatch_allows_one_raw_row_to_enter_multiple_categories(client_and_db):
    client, db = client_and_db
    file_record = UploadFileRecord(filename="multi-category.xlsx", platform="JD", row_count=1, status="done")
    db.add(file_record)
    db.flush()
    db.add_all([
        DispatchRule(
            category_code="projector",
            platform="jd",
            field="category_lv2",
            match_type="equals",
            value="平板电视",
            item_name_keyword="激光",
            priority=5,
            is_active=1,
        ),
        DispatchRule(
            category_code="tv",
            platform="jd",
            field="category_lv2",
            match_type="equals",
            value="平板电视",
            priority=10,
            is_active=1,
        ),
    ])
    db.add(RawDataRecord(
        file_id=file_record.id,
        platform="jd",
        month=202605,
        item_id="laser-tv-1",
        category_lv1="大 家 电",
        category_lv2="平板电视",
        item_name="海信激光电视 100英寸",
    ))
    db.commit()

    response = client.post("/api/dispatch/run", json={"file_id": file_record.id})

    assert response.status_code == 200
    payload = response.json()
    assert payload["total_rows"] == 1
    assert payload["dispatched_rows"] == 2
    assert payload["unmatched_rows"] == 0
    items = db.query(DispatchItem).filter_by(batch_id=payload["id"]).order_by(DispatchItem.category_code).all()
    assert [(item.category_code, item.raw_data_id) for item in items] == [
        ("projector", items[0].raw_data_id),
        ("tv", items[1].raw_data_id),
    ]


def test_run_dispatch_with_category_code_only_dispatches_target_category(client_and_db):
    client, db = client_and_db
    file_record = UploadFileRecord(filename="category-rerun.xlsx", platform="JD", row_count=3, status="done")
    db.add(file_record)
    db.flush()
    db.add_all([
        DispatchRule(
            category_code="headphone",
            platform="jd",
            field="category_lv1",
            match_type="contains",
            value="耳机",
            priority=1,
            is_active=1,
        ),
        DispatchRule(
            category_code="speaker",
            platform="jd",
            field="item_name",
            match_type="contains",
            value="音箱",
            priority=1,
            is_active=1,
        ),
    ])
    db.add_all([
        RawDataRecord(file_id=file_record.id, platform="jd", item_id="raw-1", category_lv1="耳机", item_name="旗舰耳机"),
        RawDataRecord(file_id=file_record.id, platform="jd", item_id="raw-2", category_lv1="音频", item_name="蓝牙音箱"),
        RawDataRecord(file_id=file_record.id, platform="jd", item_id="raw-3", category_lv1="耳机", item_name="耳机音箱套装"),
    ])
    db.commit()

    response = client.post("/api/dispatch/run", json={"file_id": file_record.id, "category_code": "headphone"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "done"
    assert payload["total_rows"] == 3
    assert payload["dispatched_rows"] == 2
    assert payload["unmatched_rows"] == 1
    items = db.query(DispatchItem).filter_by(batch_id=payload["id"]).order_by(DispatchItem.raw_data_id).all()
    assert [item.category_code for item in items] == ["headphone", "headphone"]
    assert {item.raw_data_id for item in items} == {
        db.query(RawDataRecord).filter_by(item_id="raw-1").one().id,
        db.query(RawDataRecord).filter_by(item_id="raw-3").one().id,
    }


def test_run_dispatch_with_category_code_preserves_other_categories_from_latest_batch(client_and_db):
    client, db = client_and_db
    file_record = UploadFileRecord(filename="category-merge.xlsx", platform="JD", row_count=3, status="done")
    db.add(file_record)
    db.flush()
    db.add_all([
        DispatchRule(
            category_code="headphone",
            platform="jd",
            field="item_name",
            match_type="contains",
            value="耳机",
            priority=1,
            is_active=1,
        ),
        DispatchRule(
            category_code="speaker",
            platform="jd",
            field="item_name",
            match_type="contains",
            value="音箱",
            priority=1,
            is_active=1,
        ),
    ])
    db.flush()
    rows = [
        RawDataRecord(file_id=file_record.id, platform="JD", item_id="raw-1", item_name="降噪耳机"),
        RawDataRecord(file_id=file_record.id, platform="JD", item_id="raw-2", item_name="蓝牙音箱"),
        RawDataRecord(file_id=file_record.id, platform="JD", item_id="raw-3", item_name="普通商品"),
    ]
    db.add_all(rows)
    db.commit()

    full_response = client.post("/api/dispatch/run", json={"file_id": file_record.id})
    assert full_response.status_code == 200
    full_batch_id = full_response.json()["id"]
    speaker_row_id = db.query(RawDataRecord).filter_by(item_id="raw-2").one().id
    assert db.query(DispatchItem).filter_by(batch_id=full_batch_id, category_code="speaker", raw_data_id=speaker_row_id).count() == 1

    scoped_response = client.post("/api/dispatch/run", json={"file_id": file_record.id, "category_code": "headphone"})

    assert scoped_response.status_code == 200
    payload = scoped_response.json()
    assert payload["total_rows"] == 3
    assert payload["dispatched_rows"] == 2
    assert payload["unmatched_rows"] == 1
    items = db.query(DispatchItem).filter_by(batch_id=payload["id"]).order_by(DispatchItem.category_code).all()
    assert [(item.category_code, item.raw_data_id) for item in items] == [
        ("headphone", db.query(RawDataRecord).filter_by(item_id="raw-1").one().id),
        ("speaker", speaker_row_id),
    ]


def test_run_dispatch_with_category_code_requires_active_rule(client_and_db):
    client, db = client_and_db
    file_record = UploadFileRecord(filename="no-rule.xlsx", platform="JD", row_count=1, status="done")
    db.add(file_record)
    db.flush()
    db.add(RawDataRecord(file_id=file_record.id, platform="jd", item_id="raw-1", category_lv1="耳机", item_name="旗舰耳机"))
    db.commit()

    response = client.post("/api/dispatch/run", json={"file_id": file_record.id, "category_code": "headphone"})

    assert response.status_code == 400
    assert response.json()["detail"] == "该品类没有可用分发规则"
    assert db.query(DispatchBatch).count() == 0


def test_run_dispatch_deduplicates_same_category_by_priority_then_id(client_and_db):
    client, db = client_and_db
    file_record = UploadFileRecord(filename="same-category.xlsx", platform="JD", row_count=1, status="done")
    db.add(file_record)
    db.flush()
    lower_priority_rule = DispatchRule(
        category_code="projector",
        platform="jd",
        field="category_lv2",
        match_type="equals",
        value="平板电视",
        item_name_keyword="激光",
        priority=5,
        is_active=1,
    )
    higher_priority_rule = DispatchRule(
        category_code="projector",
        platform="jd",
        field="item_name",
        match_type="contains",
        value="激光电视",
        priority=20,
        is_active=1,
    )
    db.add_all([lower_priority_rule, higher_priority_rule])
    db.flush()
    db.add(RawDataRecord(
        file_id=file_record.id,
        platform="jd",
        month=202605,
        item_id="laser-tv-2",
        category_lv2="平板电视",
        item_name="海信激光电视 100英寸",
    ))
    db.commit()

    response = client.post("/api/dispatch/run", json={"file_id": file_record.id})

    assert response.status_code == 200
    payload = response.json()
    assert payload["dispatched_rows"] == 1
    assert payload["unmatched_rows"] == 0
    items = db.query(DispatchItem).filter_by(batch_id=payload["id"]).all()
    assert len(items) == 1
    assert items[0].category_code == "projector"
    assert items[0].matched_rule_id == lower_priority_rule.id



def test_run_dispatch_rule_value_matches_any_split_value(client_and_db):
    client, db = client_and_db
    file_record = UploadFileRecord(filename="value-split.xlsx", platform="JD", row_count=3, status="done")
    db.add(file_record)
    db.flush()
    db.add(DispatchRule(
        category_code="soundbar",
        platform="jd",
        field="category_lv2",
        match_type="equals",
        value="回音壁/Soundbar音响、条形音箱",
        priority=1,
        is_active=1,
    ))
    for idx, category_lv2 in enumerate(["回音壁", "Soundbar音响", "家庭影院"], start=1):
        db.add(RawDataRecord(
            file_id=file_record.id,
            platform="jd",
            month=202605,
            item_id=f"item-{idx}",
            category_lv2=category_lv2,
            item_name=f"商品 {idx}",
        ))
    db.commit()

    response = client.post("/api/dispatch/run", json={"file_id": file_record.id})

    assert response.status_code == 200
    payload = response.json()
    assert payload["dispatched_rows"] == 2
    assert payload["unmatched_rows"] == 1
    items = db.query(DispatchItem).filter_by(batch_id=payload["id"]).all()
    assert [item.raw_data_id for item in items] == [1, 2]



def test_run_dispatch_rule_value_still_matches_full_value_with_slash(client_and_db):
    client, db = client_and_db
    file_record = UploadFileRecord(filename="value-full-slash.xlsx", platform="TMALL", row_count=1, status="done")
    db.add(file_record)
    db.flush()
    db.add(DispatchRule(
        category_code="tablet",
        platform="tmall",
        field="category_lv1",
        match_type="equals",
        value="平板电脑/MID",
        priority=1,
        is_active=1,
    ))
    db.add(RawDataRecord(
        file_id=file_record.id,
        platform="tmall",
        month=202605,
        item_id="item-1",
        category_lv1="平板电脑/MID",
        item_name="平板商品",
    ))
    db.commit()

    response = client.post("/api/dispatch/run", json={"file_id": file_record.id})

    assert response.status_code == 200
    payload = response.json()
    assert payload["dispatched_rows"] == 1
    assert payload["unmatched_rows"] == 0



def test_run_dispatch_item_name_keyword_matches_any_split_keyword(client_and_db):
    client, db = client_and_db
    file_record = UploadFileRecord(filename="keyword.xlsx", platform="JD", row_count=5, status="done")
    db.add(file_record)
    db.flush()
    db.add(DispatchRule(
        category_code="headphone",
        platform="jd",
        field="category_lv1",
        match_type="contains",
        value="耳机",
        item_name_keyword="旗舰, Pro，Ultra、礼盒\nMax",
        priority=1,
        is_active=1,
    ))
    for idx, item_name in enumerate(["标准款", "Pro版", "Ultra版", "Max 版", "旗舰款"], start=1):
        db.add(RawDataRecord(
            file_id=file_record.id,
            platform="jd",
            month=202605,
            item_id=f"item-{idx}",
            category_lv1="蓝牙耳机",
            item_name=item_name,
        ))
    db.commit()

    response = client.post("/api/dispatch/run", json={"file_id": file_record.id})

    assert response.status_code == 200
    payload = response.json()
    assert payload["dispatched_rows"] == 4
    assert payload["unmatched_rows"] == 1
    items = db.query(DispatchItem).filter_by(batch_id=payload["id"]).all()
    assert [item.raw_data_id for item in items] == [2, 3, 4, 5]


def test_run_dispatch_marks_error_and_rolls_back_partial_items(client_and_db, monkeypatch):
    client, db = client_and_db
    file_record = UploadFileRecord(filename="error.xlsx", platform="JD", row_count=3, status="done")
    db.add(file_record)
    db.flush()
    db.add(DispatchRule(
        category_code="headphone",
        platform="jd",
        field="item_name",
        match_type="contains",
        value="商品",
        priority=1,
        is_active=1,
    ))
    for idx, item_name in enumerate(["商品 1", "商品 2", "boom"], start=1):
        db.add(RawDataRecord(
            file_id=file_record.id,
            platform="jd",
            month=202605,
            item_id=f"item-{idx}",
            item_name=item_name,
        ))
    db.commit()

    def fail_on_boom(row, rule):
        if row.item_name == "boom":
            raise RuntimeError("boom")
        return dispatch_api._field_value(row, rule.field).find(rule.value) >= 0

    monkeypatch.setattr(dispatch_api, "_rule_matches", fail_on_boom)

    with pytest.raises(RuntimeError):
        client.post("/api/dispatch/run", json={"file_id": file_record.id})

    batch = db.query(DispatchBatch).one()
    assert batch.status == "error"
    assert db.query(DispatchItem).count() == 0


def test_get_batch_stats_returns_category_names_and_rule_counts(client_and_db):
    client, db = client_and_db
    db.add(Category(code="headphone", name="耳机", sort_order=1))
    db.add(Category(code="speaker", name="音箱", sort_order=2))
    file_record = UploadFileRecord(filename="stats.xlsx", platform="JD", row_count=4, status="done")
    db.add(file_record)
    db.flush()
    raw_rows = [
        RawDataRecord(file_id=file_record.id, platform="jd", item_id="item-1", category_lv1="蓝牙耳机", item_name="普通商品"),
        RawDataRecord(file_id=file_record.id, platform="tmall", item_id="item-2", category_lv1="头戴耳机", item_name="普通商品"),
        RawDataRecord(file_id=file_record.id, platform="jd", item_id="item-3", category_lv1="其他", item_name="无线音箱"),
    ]
    db.add_all(raw_rows)
    db.flush()
    rule_one = DispatchRule(
        category_code="headphone",
        platform="jd",
        field="category_lv1",
        match_type="contains",
        value="耳机",
        item_name_keyword=None,
        priority=1,
        is_active=1,
    )
    rule_two = DispatchRule(
        category_code="speaker",
        platform=None,
        field="item_name",
        match_type="contains",
        value="音箱",
        item_name_keyword="无线",
        priority=2,
        is_active=1,
    )
    db.add(rule_one)
    db.add(rule_two)
    db.flush()
    batch = DispatchBatch(
        file_id=file_record.id,
        status="done",
        total_rows=4,
        dispatched_rows=3,
        unmatched_rows=1,
    )
    db.add(batch)
    db.flush()
    db.add_all([
        DispatchItem(batch_id=batch.id, raw_data_id=raw_rows[0].id, category_code="headphone", matched_rule_id=rule_one.id),
        DispatchItem(batch_id=batch.id, raw_data_id=raw_rows[1].id, category_code="headphone", matched_rule_id=rule_one.id),
        DispatchItem(batch_id=batch.id, raw_data_id=raw_rows[2].id, category_code="speaker", matched_rule_id=rule_two.id),
    ])
    db.commit()

    response = client.get(f"/api/dispatch/batches/{batch.id}/stats")

    assert response.status_code == 200
    payload = response.json()
    assert payload["batch_id"] == batch.id
    assert payload["total_rows"] == 4
    assert payload["dispatched_rows"] == 3
    assert payload["unmatched_rows"] == 1
    assert payload["categories"] == [
        {
            "category_code": "headphone",
            "category_name": "耳机",
            "count": 2,
            "platforms": [{"platform": "jd", "count": 1}, {"platform": "tmall", "count": 1}],
        },
        {
            "category_code": "speaker",
            "category_name": "音箱",
            "count": 1,
            "platforms": [{"platform": "jd", "count": 1}],
        },
    ]
    assert payload["rules"] == [
        {
            "rule_id": rule_one.id,
            "category_code": "headphone",
            "category_name": "耳机",
            "field": "category_lv1",
            "match_type": "contains",
            "value": "耳机",
            "item_name_keyword": None,
            "platform": "jd",
            "priority": 1,
            "is_active": 1,
            "count": 2,
            "assigned_count": 2,
        },
        {
            "rule_id": rule_two.id,
            "category_code": "speaker",
            "category_name": "音箱",
            "field": "item_name",
            "match_type": "contains",
            "value": "音箱",
            "item_name_keyword": "无线",
            "platform": None,
            "priority": 2,
            "is_active": 1,
            "count": 1,
            "assigned_count": 1,
        },
    ]


def test_get_batch_stats_rule_count_uses_assigned_rows_without_recount(client_and_db):
    client, db = client_and_db
    db.add(Category(code="projector", name="投影仪", sort_order=1))
    file_record = UploadFileRecord(filename="actual-rule-count.xlsx", platform="DY", row_count=31, status="done")
    db.add(file_record)
    db.flush()
    high_priority_rule = DispatchRule(
        category_code="projector",
        platform="dy",
        field="item_name",
        match_type="contains",
        value="激光",
        item_name_keyword="电视",
        priority=5,
        is_active=1,
    )
    low_priority_rule = DispatchRule(
        category_code="projector",
        platform="dy",
        field="category_lv3",
        match_type="equals",
        value="激光电视",
        item_name_keyword=None,
        priority=10,
        is_active=1,
    )
    db.add_all([high_priority_rule, low_priority_rule])
    db.flush()
    raw_rows = []
    for idx in range(31):
        raw_rows.append(RawDataRecord(
            file_id=file_record.id,
            platform="dy",
            category_lv3="激光电视",
            item_name="激光电视" if idx < 28 else "普通商品",
        ))
    db.add_all(raw_rows)
    db.flush()
    batch = DispatchBatch(file_id=file_record.id, status="done", total_rows=31, dispatched_rows=31, unmatched_rows=0)
    db.add(batch)
    db.flush()
    db.add_all([
        *[
            DispatchItem(
                batch_id=batch.id,
                raw_data_id=raw_rows[idx].id,
                category_code="projector",
                matched_rule_id=high_priority_rule.id,
            )
            for idx in range(28)
        ],
        *[
            DispatchItem(
                batch_id=batch.id,
                raw_data_id=raw_rows[idx].id,
                category_code="projector",
                matched_rule_id=low_priority_rule.id,
            )
            for idx in range(28, 31)
        ],
    ])
    db.commit()

    response = client.get(f"/api/dispatch/batches/{batch.id}/stats")

    assert response.status_code == 200
    rules = {row["rule_id"]: row for row in response.json()["rules"]}
    assert rules[low_priority_rule.id]["count"] == 3
    assert rules[low_priority_rule.id]["assigned_count"] == 3


def test_get_batch_stats_returns_404_for_missing_batch(client_and_db):
    client, _ = client_and_db

    response = client.get("/api/dispatch/batches/999/stats")

    assert response.status_code == 404


def test_get_batch_stats_preserves_deleted_rule_counts(client_and_db):
    client, db = client_and_db
    db.add(Category(code="headphone", name="耳机", sort_order=1))
    file_record = UploadFileRecord(filename="deleted-rule.xlsx", platform="JD", row_count=1, status="done")
    db.add(file_record)
    db.flush()
    batch = DispatchBatch(
        file_id=file_record.id,
        status="done",
        total_rows=1,
        dispatched_rows=1,
        unmatched_rows=0,
    )
    db.add(batch)
    db.flush()
    deleted_rule_id = 12345
    db.add(DispatchItem(
        batch_id=batch.id,
        raw_data_id=1,
        category_code="headphone",
        matched_rule_id=deleted_rule_id,
    ))
    db.commit()

    response = client.get(f"/api/dispatch/batches/{batch.id}/stats")

    assert response.status_code == 200
    payload = response.json()
    assert payload["rules"] == [
        {
            "rule_id": deleted_rule_id,
            "category_code": "headphone",
            "category_name": "耳机",
            "field": None,
            "match_type": None,
            "value": None,
            "item_name_keyword": None,
            "platform": None,
            "priority": None,
            "is_active": None,
            "count": 1,
            "assigned_count": 1,
        },
    ]


def test_get_batch_stats_groups_deleted_rule_counts_by_rule_id(client_and_db):
    client, db = client_and_db
    db.add(Category(code="headphone", name="耳机", sort_order=1))
    db.add(Category(code="speaker", name="音箱", sort_order=2))
    file_record = UploadFileRecord(filename="deleted-rule-multi-category.xlsx", platform="JD", row_count=2, status="done")
    db.add(file_record)
    db.flush()
    batch = DispatchBatch(
        file_id=file_record.id,
        status="done",
        total_rows=2,
        dispatched_rows=2,
        unmatched_rows=0,
    )
    db.add(batch)
    db.flush()
    deleted_rule_id = 12345
    db.add_all([
        DispatchItem(batch_id=batch.id, raw_data_id=1, category_code="speaker", matched_rule_id=deleted_rule_id),
        DispatchItem(batch_id=batch.id, raw_data_id=2, category_code="headphone", matched_rule_id=deleted_rule_id),
    ])
    db.commit()

    response = client.get(f"/api/dispatch/batches/{batch.id}/stats")

    assert response.status_code == 200
    payload = response.json()
    assert payload["rules"] == [
        {
            "rule_id": deleted_rule_id,
            "category_code": "headphone",
            "category_name": "耳机",
            "field": None,
            "match_type": None,
            "value": None,
            "item_name_keyword": None,
            "platform": None,
            "priority": None,
            "is_active": None,
            "count": 2,
            "assigned_count": 2,
        },
    ]


def test_get_batch_unmatched_excludes_dispatched_rows_and_paginates(client_and_db):
    client, db = client_and_db
    file_record = UploadFileRecord(filename="unmatched.xlsx", platform="JD", row_count=5, status="done")
    other_file = UploadFileRecord(filename="other.xlsx", platform="JD", row_count=1, status="done")
    db.add_all([file_record, other_file])
    db.flush()
    rows = []
    for idx in range(1, 6):
        row = RawDataRecord(
            file_id=file_record.id,
            platform="jd",
            month=202605,
            item_id=f"item-{idx}",
            item_name=f"商品 {idx}",
            category_lv1="一级",
            category_lv2="二级",
            category_lv3="三级",
            brand_raw=f"品牌 {idx}",
            shop_name=f"店铺 {idx}",
            price=idx * 10,
            sales_qty=idx,
            sales_amount=idx * 100,
        )
        db.add(row)
        rows.append(row)
    other_row = RawDataRecord(file_id=other_file.id, item_id="other", item_name="其他文件")
    db.add(other_row)
    db.flush()
    batch = DispatchBatch(file_id=file_record.id, status="done", total_rows=5, dispatched_rows=2, unmatched_rows=3)
    other_batch = DispatchBatch(file_id=file_record.id, status="done")
    db.add_all([batch, other_batch])
    db.flush()
    db.add_all([
        DispatchItem(batch_id=batch.id, raw_data_id=rows[1].id, category_code="headphone"),
        DispatchItem(batch_id=batch.id, raw_data_id=rows[3].id, category_code="speaker"),
        DispatchItem(batch_id=batch.id, raw_data_id=None, category_code="unknown"),
        DispatchItem(batch_id=other_batch.id, raw_data_id=rows[0].id, category_code="other-batch"),
    ])
    db.commit()

    response = client.get(f"/api/dispatch/batches/{batch.id}/unmatched", params={"page": 2, "page_size": 2})

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 3
    assert payload["page"] == 2
    assert payload["page_size"] == 2
    assert [item["item_id"] for item in payload["items"]] == ["item-5"]
    assert payload["items"][0] == {
        "id": rows[4].id,
        "item_id": "item-5",
        "item_name": "商品 5",
        "platform": "jd",
        "month": 202605,
        "category_lv1": "一级",
        "category_lv2": "二级",
        "category_lv3": "三级",
        "brand_raw": "品牌 5",
        "shop_name": "店铺 5",
        "price": 50.0,
        "sales_qty": 5,
        "sales_amount": 500.0,
    }


def test_get_batch_unmatched_filters_by_item_id_or_item_name_keyword(client_and_db):
    client, db = client_and_db
    file_record = UploadFileRecord(filename="keyword.xlsx", platform="JD", row_count=4, status="done")
    db.add(file_record)
    db.flush()
    rows = [
        RawDataRecord(file_id=file_record.id, item_id="SKU-ALPHA", item_name="普通商品"),
        RawDataRecord(file_id=file_record.id, item_id="SKU-BETA", item_name="降噪耳机"),
        RawDataRecord(file_id=file_record.id, item_id="SKU-GAMMA", item_name="无线音箱"),
        RawDataRecord(file_id=file_record.id, item_id="SKU-DELTA", item_name="已分发耳机"),
    ]
    db.add_all(rows)
    db.flush()
    batch = DispatchBatch(file_id=file_record.id, status="done", total_rows=4, dispatched_rows=1, unmatched_rows=3)
    db.add(batch)
    db.flush()
    db.add(DispatchItem(batch_id=batch.id, raw_data_id=rows[3].id, category_code="headphone"))
    db.commit()

    response = client.get(f"/api/dispatch/batches/{batch.id}/unmatched", params={"keyword": "耳机"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 1
    assert [item["item_id"] for item in payload["items"]] == ["SKU-BETA"]


def test_get_batch_unmatched_returns_404_for_missing_batch(client_and_db):
    client, _ = client_and_db

    response = client.get("/api/dispatch/batches/999/unmatched")

    assert response.status_code == 404
    assert response.json()["detail"] == "批次不存在"


def test_get_batch_unmatched_returns_empty_when_batch_has_no_file(client_and_db):
    client, db = client_and_db
    batch = DispatchBatch(file_id=None, status="done")
    db.add(batch)
    db.commit()

    response = client.get(f"/api/dispatch/batches/{batch.id}/unmatched")

    assert response.status_code == 200
    assert response.json() == {"total": 0, "page": 1, "page_size": 20, "items": []}


def _read_workbook(response):
    return openpyxl.load_workbook(io.BytesIO(response.content))


def _sheet_rows(sheet):
    return list(sheet.iter_rows(values_only=True))


def test_export_batch_raw_data_filters_by_category_and_platform(client_and_db):
    client, db = client_and_db
    template = ColumnTemplate(
        name="京东模板",
        module="sales",
        mapping={"平台": "platform", "宝贝ID": "item_id", "宝贝名称": "item_name"},
        ignore_columns=[],
    )
    db.add(template)
    db.flush()
    file_record = UploadFileRecord(filename="export.xlsx", platform="JD", row_count=3, status="done", template_id=template.id)
    db.add(file_record)
    db.flush()
    rows = [
        RawDataRecord(file_id=file_record.id, platform="jd", item_id="jd-1", item_name="京东商品1"),
        RawDataRecord(file_id=file_record.id, platform="jd", item_id="jd-2", item_name="京东商品2"),
        RawDataRecord(file_id=file_record.id, platform="tmall", item_id="tm-1", item_name="天猫商品"),
    ]
    db.add_all(rows)
    db.flush()
    batch = DispatchBatch(file_id=file_record.id, status="done", total_rows=3, dispatched_rows=3, unmatched_rows=0)
    db.add(batch)
    db.flush()
    db.add_all([
        DispatchItem(batch_id=batch.id, raw_data_id=rows[0].id, category_code="headphone"),
        DispatchItem(batch_id=batch.id, raw_data_id=rows[1].id, category_code="headphone"),
        DispatchItem(batch_id=batch.id, raw_data_id=rows[2].id, category_code="headphone"),
    ])
    db.commit()

    response = client.get(
        f"/api/dispatch/batches/{batch.id}/export",
        params={"category_code": "headphone", "platform": "jd"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = _read_workbook(response)
    rows = _sheet_rows(workbook.active)
    assert rows[0][:3] == ("平台", "宝贝ID", "宝贝名称")
    assert [row[1] for row in rows[1:]] == ["jd-1", "jd-2"]


def test_export_batch_raw_data_restores_template_columns_and_extra_data(client_and_db):
    client, db = client_and_db
    template = ColumnTemplate(
        name="扩展模板",
        module="sales",
        mapping={
            "平台": "platform",
            "月份": "month",
            "宝贝ID": "item_id",
            "宝贝名称": "item_name",
            "自定义字段": "__ext__",
        },
        ignore_columns=[],
    )
    db.add(template)
    db.flush()
    file_record = UploadFileRecord(filename="template.xlsx", platform="JD", row_count=1, status="done", template_id=template.id)
    db.add(file_record)
    db.flush()
    raw = RawDataRecord(
        file_id=file_record.id,
        platform="jd",
        month=202605,
        item_id="sku-1",
        item_name="模板商品",
        extra_data={"自定义字段": "扩展值", "额外保留": "保留值"},
    )
    db.add(raw)
    db.flush()
    batch = DispatchBatch(file_id=file_record.id, status="done", total_rows=1, dispatched_rows=1, unmatched_rows=0)
    db.add(batch)
    db.flush()
    db.add(DispatchItem(batch_id=batch.id, raw_data_id=raw.id, category_code="headphone"))
    db.commit()

    response = client.get(f"/api/dispatch/batches/{batch.id}/export", params={"category_code": "headphone"})

    assert response.status_code == 200
    rows = _sheet_rows(_read_workbook(response).active)
    assert rows[0] == ("平台", "月份", "宝贝ID", "宝贝名称", "自定义字段", "额外保留", "_raw_data_id", "_source_filename")
    assert rows[1][:6] == ("jd", 202605, "sku-1", "模板商品", "扩展值", "保留值")


def test_export_batch_raw_data_splits_multiple_templates_into_sheets(client_and_db):
    client, db = client_and_db
    template_one = ColumnTemplate(name="京东模板", module="sales", mapping={"京东ID": "item_id"}, ignore_columns=[])
    template_two = ColumnTemplate(name="天猫模板", module="sales", mapping={"天猫ID": "item_id"}, ignore_columns=[])
    db.add_all([template_one, template_two])
    db.flush()
    file_one = UploadFileRecord(filename="jd.xlsx", platform="JD", row_count=1, status="done", template_id=template_one.id)
    file_two = UploadFileRecord(filename="tmall.xlsx", platform="TMALL", row_count=1, status="done", template_id=template_two.id)
    db.add_all([file_one, file_two])
    db.flush()
    raw_one = RawDataRecord(file_id=file_one.id, platform="jd", item_id="jd-1")
    raw_two = RawDataRecord(file_id=file_two.id, platform="tmall", item_id="tm-1")
    db.add_all([raw_one, raw_two])
    db.flush()
    batch = DispatchBatch(file_id=file_one.id, status="done", total_rows=2, dispatched_rows=2, unmatched_rows=0)
    db.add(batch)
    db.flush()
    db.add_all([
        DispatchItem(batch_id=batch.id, raw_data_id=raw_one.id, category_code="headphone"),
        DispatchItem(batch_id=batch.id, raw_data_id=raw_two.id, category_code="headphone"),
    ])
    db.commit()

    response = client.get(f"/api/dispatch/batches/{batch.id}/export", params={"category_code": "headphone"})

    assert response.status_code == 200
    workbook = _read_workbook(response)
    assert len(workbook.sheetnames) == 2
    headers = [workbook[sheet_name][1][0].value for sheet_name in workbook.sheetnames]
    assert headers == ["京东ID", "天猫ID"]


def test_create_dispatch_export_job_requires_category_platform_or_month(client_and_db):
    client, _ = client_and_db

    response = client.post("/api/dispatch/export", json={})

    assert response.status_code == 400
    assert response.json()["detail"] == "请选择品类、平台或月份后再导出"


def test_create_dispatch_export_job_rejects_invalid_month(client_and_db):
    client, _ = client_and_db

    response = client.post("/api/dispatch/export", json={"month": 202613})

    assert response.status_code == 400
    assert response.json()["detail"] == "月份格式应为 YYYYMM"


def test_create_dispatch_export_job_rejects_empty_result_before_creating_job(client_and_db):
    client, db = client_and_db

    response = client.post("/api/dispatch/export", json={"category_code": "camera", "platform": "jd", "month": 202601})

    assert response.status_code == 400
    assert response.json()["detail"] == "当前筛选条件无可导出数据，请调整月份、品类或平台"
    assert db.query(WorkbenchExportJob).count() == 0


def test_create_dispatch_export_job_rejects_when_exports_are_busy(client_and_db):
    client, db = client_and_db
    file_record = UploadFileRecord(filename="busy.xlsx", platform="JD", row_count=1, status="done")
    db.add(file_record)
    db.flush()
    raw = RawDataRecord(file_id=file_record.id, platform="jd", item_id="busy-row")
    db.add(raw)
    db.flush()
    batch = DispatchBatch(file_id=file_record.id, status="done", total_rows=1, dispatched_rows=1, unmatched_rows=0)
    db.add(batch)
    db.flush()
    db.add(DispatchItem(batch_id=batch.id, raw_data_id=raw.id, category_code="headphone"))
    db.add_all([
        WorkbenchExportJob(status="running", progress=10),
        WorkbenchExportJob(status="pending", progress=0),
    ])
    db.commit()

    response = client.post("/api/dispatch/export", json={"category_code": "headphone"})

    assert response.status_code == 429
    assert response.json()["detail"] == "当前导出任务较多，请稍后再试"


def test_list_dispatch_export_jobs_includes_filters_and_download_url(client_and_db):
    client, db = client_and_db
    older_job = WorkbenchExportJob(
        status="running",
        progress=35,
        category_code="camera",
        platform="jd",
        month=202605,
    )
    done_job = WorkbenchExportJob(
        status="done",
        progress=100,
        category_code="headphone",
        platform="tmall",
        month=202606,
        file_token="download-token",
        filename="分发结果_耳机_天猫_202606.xlsx",
    )
    db.add_all([older_job, done_job])
    db.commit()

    response = client.get("/api/dispatch/export/jobs")

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 2
    assert [item["job_id"] for item in payload["items"]] == [done_job.id, older_job.id]
    latest = payload["items"][0]
    assert latest["status"] == "done"
    assert latest["progress"] == 100
    assert latest["category_code"] == "headphone"
    assert latest["platform"] == "tmall"
    assert latest["month"] == 202606
    assert latest["filename"] == "分发结果_耳机_天猫_202606.xlsx"
    assert latest["download_url"] == "/api/dispatch/export/download/download-token"
    assert latest["created_at"]


def test_dispatch_export_job_uses_latest_done_batch_per_file(client_and_db):
    client, db = client_and_db
    template = ColumnTemplate(name="下载模板", module="sales", mapping={"商品ID": "item_id"}, ignore_columns=[])
    db.add(template)
    db.flush()
    file_record = UploadFileRecord(filename="latest.xlsx", platform="JD", row_count=2, status="done", template_id=template.id)
    other_file = UploadFileRecord(filename="other.xlsx", platform="JD", row_count=1, status="done", template_id=template.id)
    db.add_all([file_record, other_file])
    db.flush()
    old_raw = RawDataRecord(file_id=file_record.id, platform="jd", item_id="old-row")
    latest_raw = RawDataRecord(file_id=file_record.id, platform="jd", item_id="latest-row")
    other_raw = RawDataRecord(file_id=other_file.id, platform="tmall", item_id="other-row")
    db.add_all([old_raw, latest_raw, other_raw])
    db.flush()
    old_batch = DispatchBatch(file_id=file_record.id, status="done", total_rows=2, dispatched_rows=1, unmatched_rows=1)
    latest_batch = DispatchBatch(file_id=file_record.id, status="done", total_rows=2, dispatched_rows=1, unmatched_rows=1)
    other_batch = DispatchBatch(file_id=other_file.id, status="done", total_rows=1, dispatched_rows=1, unmatched_rows=0)
    db.add_all([old_batch, latest_batch, other_batch])
    db.flush()
    db.add_all([
        DispatchItem(batch_id=old_batch.id, raw_data_id=old_raw.id, category_code="headphone"),
        DispatchItem(batch_id=latest_batch.id, raw_data_id=latest_raw.id, category_code="headphone"),
        DispatchItem(batch_id=other_batch.id, raw_data_id=other_raw.id, category_code="headphone"),
    ])
    job = WorkbenchExportJob(status="pending", progress=0)
    db.add(job)
    db.commit()

    dispatch_api._run_dispatch_export_thread(job.id, {"category_code": "headphone", "platform": None})

    db.refresh(job)
    assert job.status == "done"
    response = client.get(f"/api/dispatch/export/download/{job.file_token}")
    assert response.status_code == 200
    rows = _sheet_rows(_read_workbook(response).active)
    assert [row[0] for row in rows[1:]] == ["latest-row", "other-row"]


def test_dispatch_export_job_reads_and_writes_in_pages(client_and_db, monkeypatch):
    client, db = client_and_db
    monkeypatch.setattr(dispatch_api, "DISPATCH_PAGE_SIZE", 1)
    template = ColumnTemplate(name="分页模板", module="sales", mapping={"商品ID": "item_id"}, ignore_columns=[])
    db.add(template)
    db.flush()
    file_record = UploadFileRecord(filename="paged.xlsx", platform="JD", row_count=2, status="done", template_id=template.id)
    db.add(file_record)
    db.flush()
    raw_one = RawDataRecord(file_id=file_record.id, platform="jd", item_id="row-1")
    raw_two = RawDataRecord(file_id=file_record.id, platform="jd", item_id="row-2")
    db.add_all([raw_one, raw_two])
    db.flush()
    batch = DispatchBatch(file_id=file_record.id, status="done", total_rows=2, dispatched_rows=2, unmatched_rows=0)
    db.add(batch)
    db.flush()
    db.add_all([
        DispatchItem(batch_id=batch.id, raw_data_id=raw_one.id, category_code="headphone"),
        DispatchItem(batch_id=batch.id, raw_data_id=raw_two.id, category_code="headphone"),
    ])
    job = WorkbenchExportJob(status="pending", progress=0)
    db.add(job)
    db.commit()

    dispatch_api._run_dispatch_export_thread(job.id, {"category_code": "headphone", "platform": None})

    db.refresh(job)
    assert job.status == "done"
    response = client.get(f"/api/dispatch/export/download/{job.file_token}")
    assert response.status_code == 200
    rows = _sheet_rows(_read_workbook(response).active)
    assert [row[0] for row in rows[1:]] == ["row-1", "row-2"]


def test_dispatch_export_job_filters_by_platform(client_and_db):
    client, db = client_and_db
    template = ColumnTemplate(name="平台模板", module="sales", mapping={"商品ID": "item_id", "平台": "platform"}, ignore_columns=[])
    db.add(template)
    db.flush()
    file_record = UploadFileRecord(filename="platform.xlsx", platform="JD", row_count=2, status="done", template_id=template.id)
    db.add(file_record)
    db.flush()
    jd_raw = RawDataRecord(file_id=file_record.id, platform="jd", item_id="jd-row")
    tmall_raw = RawDataRecord(file_id=file_record.id, platform="tmall", item_id="tmall-row")
    db.add_all([jd_raw, tmall_raw])
    db.flush()
    batch = DispatchBatch(file_id=file_record.id, status="done", total_rows=2, dispatched_rows=2, unmatched_rows=0)
    db.add(batch)
    db.flush()
    db.add_all([
        DispatchItem(batch_id=batch.id, raw_data_id=jd_raw.id, category_code="headphone"),
        DispatchItem(batch_id=batch.id, raw_data_id=tmall_raw.id, category_code="headphone"),
    ])
    job = WorkbenchExportJob(status="pending", progress=0)
    db.add(job)
    db.commit()

    dispatch_api._run_dispatch_export_thread(job.id, {"category_code": "headphone", "platform": "jd"})

    db.refresh(job)
    assert job.status == "done"
    response = client.get(f"/api/dispatch/export/download/{job.file_token}")
    assert response.status_code == 200
    rows = _sheet_rows(_read_workbook(response).active)
    assert [row[0] for row in rows[1:]] == ["jd-row"]


def test_dispatch_export_job_filters_by_raw_data_month(client_and_db):
    client, db = client_and_db
    template = ColumnTemplate(name="月份模板", module="sales", mapping={"商品ID": "item_id", "月份": "month"}, ignore_columns=[])
    db.add(template)
    db.flush()
    file_record = UploadFileRecord(filename="month.xlsx", platform="JD", row_count=2, status="done", template_id=template.id)
    db.add(file_record)
    db.flush()
    current_month_raw = RawDataRecord(file_id=file_record.id, platform="jd", month=202512, item_id="current-month")
    previous_month_raw = RawDataRecord(file_id=file_record.id, platform="jd", month=202511, item_id="previous-month")
    db.add_all([current_month_raw, previous_month_raw])
    db.flush()
    batch = DispatchBatch(file_id=file_record.id, status="done", total_rows=2, dispatched_rows=2, unmatched_rows=0)
    db.add(batch)
    db.flush()
    db.add_all([
        DispatchItem(batch_id=batch.id, raw_data_id=current_month_raw.id, category_code="camera"),
        DispatchItem(batch_id=batch.id, raw_data_id=previous_month_raw.id, category_code="camera"),
    ])
    job = WorkbenchExportJob(status="pending", progress=0)
    db.add(job)
    db.commit()

    dispatch_api._run_dispatch_export_thread(job.id, {"category_code": None, "platform": None, "month": 202512})

    db.refresh(job)
    assert job.status == "done"
    assert job.filename == "分发结果_全部品类_全部平台_202512.xlsx"
    response = client.get(f"/api/dispatch/export/download/{job.file_token}")
    assert response.status_code == 200
    rows = _sheet_rows(_read_workbook(response).active)
    assert rows[0][:2] == ("商品ID", "月份")
    assert rows[1:] == [("current-month", 202512, current_month_raw.id, "month.xlsx")]


def test_export_batch_raw_data_rejects_unfinished_batch(client_and_db):
    client, db = client_and_db
    batch = DispatchBatch(status="running")
    db.add(batch)
    db.commit()

    response = client.get(f"/api/dispatch/batches/{batch.id}/export", params={"category_code": "headphone"})

    assert response.status_code == 400
    assert response.json()["detail"] == "批次尚未完成，不能导出"


def test_export_batch_raw_data_returns_404_when_no_data(client_and_db):
    client, db = client_and_db
    batch = DispatchBatch(status="done")
    db.add(batch)
    db.commit()

    response = client.get(f"/api/dispatch/batches/{batch.id}/export", params={"category_code": "headphone"})

    assert response.status_code == 404
    assert response.json()["detail"] == "没有可导出的分发数据"


def test_batch_stats_uses_assigned_rule_counts_without_raw_recount(client_and_db, monkeypatch):
    client, db = client_and_db
    file_record = UploadFileRecord(filename="stats.xlsx", platform="DOUYIN", month_range="202606", row_count=1)
    category = Category(code="phone", name="手机")
    rule = DispatchRule(
        category_code="phone",
        field="category_lv1",
        match_type="contains",
        value="手机",
        priority=10,
        is_active=1,
    )
    db.add_all([file_record, category, rule])
    db.flush()
    batch = DispatchBatch(file_id=file_record.id, status="done", total_rows=1, dispatched_rows=1, unmatched_rows=0)
    raw = RawDataRecord(file_id=file_record.id, platform="DOUYIN", category_lv1="手机", item_name="测试手机")
    db.add_all([batch, raw])
    db.flush()
    db.add(DispatchItem(batch_id=batch.id, raw_data_id=raw.id, category_code="phone", matched_rule_id=rule.id))
    db.commit()

    def fail_if_recount_called(*args, **kwargs):
        raise AssertionError("stats endpoint should not rescan raw_data to recount rule matches")

    monkeypatch.setattr(dispatch_api, "_count_rule_matches_for_batch", fail_if_recount_called)

    response = client.get(f"/api/dispatch/batches/{batch.id}/stats")

    assert response.status_code == 200
    assert response.json()["rules"] == [
        {
            "rule_id": rule.id,
            "category_code": "phone",
            "category_name": "手机",
            "field": "category_lv1",
            "match_type": "contains",
            "value": "手机",
            "item_name_keyword": None,
            "platform": None,
            "priority": 10,
            "is_active": 1,
            "count": 1,
            "assigned_count": 1,
        }
    ]


def test_enqueue_dispatch_category_for_clean_returns_counts(client_and_db):
    client, db = client_and_db
    batch = DispatchBatch(status="done", total_rows=0, dispatched_rows=0, unmatched_rows=0)
    db.add(batch)
    db.commit()

    response = client.post(f"/api/dispatch/batches/{batch.id}/categories/headphone/enqueue-clean")

    assert response.status_code == 200
    assert response.json() == {
        "dispatch_batch_id": batch.id,
        "category_code": "headphone",
        "dispatch_count": 0,
        "pending_count": 0,
        "queued_count": 0,
    }


def test_dispatch_batches_stats_and_unmatched_are_scoped_to_visible_categories(client_and_db):
    client, db = client_and_db
    _set_dispatch_user(client, category_permissions=["headphone"])
    db.add_all([
        Category(code="headphone", name="耳机", sort_order=1),
        Category(code="speaker", name="音箱", sort_order=2),
    ])
    db.flush()

    mixed_file = UploadFileRecord(filename="mixed.xlsx", platform="JD", row_count=3, status="done")
    speaker_file = UploadFileRecord(filename="speaker.xlsx", platform="JD", row_count=1, status="done")
    db.add_all([mixed_file, speaker_file])
    db.flush()

    mixed_rows = [
        RawDataRecord(file_id=mixed_file.id, platform="jd", item_id="h-1", item_name="耳机 1"),
        RawDataRecord(file_id=mixed_file.id, platform="jd", item_id="s-1", item_name="音箱 1"),
        RawDataRecord(file_id=mixed_file.id, platform="jd", item_id="u-1", item_name="未分类 1"),
    ]
    speaker_rows = [
        RawDataRecord(file_id=speaker_file.id, platform="jd", item_id="s-only", item_name="音箱 only"),
    ]
    db.add_all(mixed_rows + speaker_rows)
    db.flush()

    mixed_batch = DispatchBatch(file_id=mixed_file.id, status="done", total_rows=3, dispatched_rows=2, unmatched_rows=1)
    speaker_batch = DispatchBatch(file_id=speaker_file.id, status="done", total_rows=1, dispatched_rows=1, unmatched_rows=0)
    db.add_all([mixed_batch, speaker_batch])
    db.flush()

    db.add_all([
        DispatchItem(batch_id=mixed_batch.id, raw_data_id=mixed_rows[0].id, category_code="headphone"),
        DispatchItem(batch_id=mixed_batch.id, raw_data_id=mixed_rows[1].id, category_code="speaker"),
        DispatchItem(batch_id=speaker_batch.id, raw_data_id=speaker_rows[0].id, category_code="speaker"),
    ])
    db.commit()

    list_response = client.get("/api/dispatch/batches")
    assert list_response.status_code == 200
    assert [item["id"] for item in list_response.json()] == [mixed_batch.id]

    stats_response = client.get(f"/api/dispatch/batches/{mixed_batch.id}/stats")
    assert stats_response.status_code == 200
    stats_payload = stats_response.json()
    assert stats_payload["dispatched_rows"] == 1
    assert stats_payload["categories"] == [
        {
            "category_code": "headphone",
            "category_name": "耳机",
            "count": 1,
            "platforms": [{"platform": "jd", "count": 1}],
        }
    ]

    unmatched_response = client.get(f"/api/dispatch/batches/{speaker_batch.id}/unmatched")
    assert unmatched_response.status_code == 403


def test_dispatch_rule_mutations_reject_unauthorized_category_code(client_and_db):
    client, db = client_and_db
    _set_dispatch_user(client, category_permissions=["headphone"])
    db.add_all([
        Category(code="headphone", name="耳机", sort_order=1),
        Category(code="speaker", name="音箱", sort_order=2),
    ])
    db.flush()

    create_response = client.post(
        "/api/dispatch/rules",
        json={
            "category_code": "speaker",
            "platform": "jd",
            "field": "category_lv1",
            "match_type": "equals",
            "value": "音箱",
            "priority": 10,
            "is_active": True,
        },
    )
    assert create_response.status_code == 403

    rule = DispatchRule(
        category_code="headphone",
        platform="jd",
        field="category_lv1",
        match_type="equals",
        value="耳机",
        priority=10,
        is_active=1,
    )
    speaker_rule = DispatchRule(
        category_code="speaker",
        platform="jd",
        field="category_lv1",
        match_type="equals",
        value="音箱",
        priority=20,
        is_active=1,
    )
    db.add_all([rule, speaker_rule])
    db.commit()

    update_response = client.put(
        f"/api/dispatch/rules/{rule.id}",
        json={
            "category_code": "speaker",
            "platform": "jd",
            "field": "category_lv1",
            "match_type": "equals",
            "value": "耳机",
            "priority": 10,
            "is_active": True,
        },
    )
    assert update_response.status_code == 403

    delete_response = client.delete(f"/api/dispatch/rules/{speaker_rule.id}")
    assert delete_response.status_code == 403


def test_dispatch_export_jobs_are_scoped_to_visible_categories(client_and_db):
    client, db = client_and_db
    _set_dispatch_user(client, category_permissions=["headphone"])

    export_dir = Path(dispatch_api.DISPATCH_EXPORT_DIR)
    headphone_job = WorkbenchExportJob(
        status="done",
        progress=100,
        category_code="headphone",
        platform="jd",
        month=202605,
        file_token="headphone-token",
        filename="headphone.xlsx",
    )
    speaker_job = WorkbenchExportJob(
        status="done",
        progress=100,
        category_code="speaker",
        platform="jd",
        month=202605,
        file_token="speaker-token",
        filename="speaker.xlsx",
    )
    db.add_all([headphone_job, speaker_job])
    db.commit()
    (export_dir / f"speaker-token_{speaker_job.filename}").write_bytes(b"speaker export")

    create_response = client.post("/api/dispatch/export", json={"category_code": "speaker"})
    assert create_response.status_code == 403

    list_response = client.get("/api/dispatch/export/jobs")
    assert list_response.status_code == 200
    payload = list_response.json()
    assert payload["total"] == 1
    assert [item["category_code"] for item in payload["items"]] == ["headphone"]

    download_response = client.get("/api/dispatch/export/download/speaker-token")
    assert download_response.status_code == 403


# ─── 批量补分发 ──────────────────────────────────────────────

class _SyncThread:
    def __init__(self, target=None, args=(), kwargs=None, daemon=None):
        self._target = target
        self._args = args
        self._kwargs = kwargs or {}

    def start(self):
        self._target(*self._args, **self._kwargs)


def _enable_sync_threads(monkeypatch):
    monkeypatch.setattr(dispatch_api.threading, "Thread", _SyncThread)


def _make_dispatch_file(db, *, filename, platform="JD", rows=None, row_count=None):
    file_record = UploadFileRecord(filename=filename, platform=platform, row_count=row_count or len(rows or []), status="done")
    db.add(file_record)
    db.flush()
    created = []
    for idx, (category_lv1, item_name) in enumerate(rows or [], start=1):
        raw = RawDataRecord(
            file_id=file_record.id,
            platform=platform.lower(),
            month=202605,
            item_id=f"{filename}-{idx}",
            category_lv1=category_lv1,
            item_name=item_name,
        )
        db.add(raw)
        created.append(raw)
    db.flush()
    return file_record, created


def _make_done_batch(db, file_record, *, category_code, raw_rows):
    batch = DispatchBatch(file_id=file_record.id, status="done", total_rows=len(raw_rows), dispatched_rows=len(raw_rows), unmatched_rows=0)
    db.add(batch)
    db.flush()
    for raw in raw_rows:
        db.add(DispatchItem(batch_id=batch.id, raw_data_id=raw.id, category_code=category_code))
    db.commit()
    return batch


def test_create_redispatch_job_validation_errors(client_and_db):
    client, db = client_and_db
    db.add_all([
        Category(code="headphone", name="耳机", sort_order=1),
        Category(code="speaker", name="音箱", sort_order=2),
    ])
    file_record, rows = _make_dispatch_file(db, filename="validation.xlsx", rows=[("手机配件", "商品 1")])
    batch = _make_done_batch(db, file_record, category_code="headphone", raw_rows=rows)

    assert client.post("/api/dispatch/redispatch", json={"batch_ids": [], "category_code": "speaker"}).status_code == 400
    assert client.post("/api/dispatch/redispatch", json={"batch_ids": [batch.id], "category_code": ""}).status_code == 400

    _set_dispatch_user(client, category_permissions=["headphone"])
    forbidden = client.post("/api/dispatch/redispatch", json={"batch_ids": [batch.id], "category_code": "speaker"})
    assert forbidden.status_code == 403

    _set_dispatch_user(client)
    missing = client.post("/api/dispatch/redispatch", json={"batch_ids": [99999], "category_code": "speaker"})
    assert missing.status_code == 404

    unfinished = DispatchBatch(file_id=file_record.id, status="running")
    db.add(unfinished)
    db.commit()
    unfinished_resp = client.post(
        "/api/dispatch/redispatch",
        json={"batch_ids": [unfinished.id], "category_code": "speaker"},
    )
    assert unfinished_resp.status_code == 400

    no_rule = client.post("/api/dispatch/redispatch", json={"batch_ids": [batch.id], "category_code": "speaker"})
    assert no_rule.status_code == 400


def test_redispatch_dispatches_target_category_and_preserves_others(client_and_db, monkeypatch):
    client, db = client_and_db
    _enable_sync_threads(monkeypatch)
    db.add_all([
        Category(code="headphone", name="耳机", sort_order=1),
        Category(code="speaker", name="音箱", sort_order=2),
    ])
    db.add(DispatchRule(
        category_code="headphone", platform="jd", field="category_lv1",
        match_type="contains", value="耳机", priority=1, is_active=1,
    ))
    db.add(DispatchRule(
        category_code="speaker", platform="jd", field="item_name",
        match_type="contains", value="音箱", priority=1, is_active=1,
    ))
    db.commit()

    file_record, rows = _make_dispatch_file(
        db,
        filename="history.xlsx",
        rows=[("蓝牙耳机", "商品 A"), ("手机配件", "无线音箱"), ("手机配件", "普通商品")],
    )
    old_batch = _make_done_batch(db, file_record, category_code="headphone", raw_rows=[rows[0]])

    response = client.post(
        "/api/dispatch/redispatch",
        json={"batch_ids": [old_batch.id], "category_code": "speaker", "skip_contained": False},
    )
    assert response.status_code == 202
    job_id = response.json()["job_id"]

    job = db.query(dispatch_api.DispatchRedispatchJob).filter_by(id=job_id).one()
    assert job.status == "done"
    assert job.success_batches == 1
    assert job.failed_batches == 0
    assert job.total_batches == 1

    item = db.query(dispatch_api.DispatchRedispatchJobItem).filter_by(job_id=job_id).one()
    assert item.status == "done"
    new_batch = db.query(DispatchBatch).filter_by(id=item.new_batch_id).one()
    new_items = db.query(DispatchItem).filter_by(batch_id=new_batch.id).all()
    assert sorted([i.category_code for i in new_items]) == ["headphone", "speaker"]
    speaker_item = next(i for i in new_items if i.category_code == "speaker")
    assert speaker_item.raw_data_id == rows[1].id


def test_redispatch_skips_file_already_containing_target_category(client_and_db, monkeypatch):
    client, db = client_and_db
    _enable_sync_threads(monkeypatch)
    db.add_all([
        Category(code="headphone", name="耳机", sort_order=1),
        Category(code="speaker", name="音箱", sort_order=2),
    ])
    db.add(DispatchRule(
        category_code="speaker", platform="jd", field="item_name",
        match_type="contains", value="音箱", priority=1, is_active=1,
    ))
    db.commit()

    file_record, rows = _make_dispatch_file(db, filename="skip.xlsx", rows=[("手机配件", "无线音箱")])
    old_batch = _make_done_batch(db, file_record, category_code="speaker", raw_rows=rows)

    response = client.post(
        "/api/dispatch/redispatch",
        json={"batch_ids": [old_batch.id], "category_code": "speaker", "skip_contained": True},
    )
    assert response.status_code == 202
    job_id = response.json()["job_id"]

    job = db.query(dispatch_api.DispatchRedispatchJob).filter_by(id=job_id).one()
    assert job.status == "done"
    assert job.success_batches == 0
    assert job.skipped_batches == 1

    item = db.query(dispatch_api.DispatchRedispatchJobItem).filter_by(job_id=job_id).one()
    assert item.status == "skipped"
    assert item.new_batch_id is None


def test_redispatch_create_rejects_when_no_rule_for_category(client_and_db):
    client, db = client_and_db
    db.add(Category(code="speaker", name="音箱", sort_order=1))
    db.commit()

    file_record, rows = _make_dispatch_file(db, filename="norulex.xlsx", rows=[("手机配件", "商品 1")])
    batch = _make_done_batch(db, file_record, category_code="headphone", raw_rows=rows)

    response = client.post(
        "/api/dispatch/redispatch",
        json={"batch_ids": [batch.id], "category_code": "speaker"},
    )
    assert response.status_code == 400


def test_redispatch_thread_item_errors_when_batch_has_no_file(client_and_db, monkeypatch):
    client, db = client_and_db
    _enable_sync_threads(monkeypatch)
    db.add(Category(code="speaker", name="音箱", sort_order=1))
    db.add(DispatchRule(
        category_code="speaker", platform=None, field="item_name",
        match_type="contains", value="音箱", priority=1, is_active=1,
    ))
    db.commit()
    batch = DispatchBatch(file_id=None, status="done", total_rows=0, dispatched_rows=0, unmatched_rows=0)
    db.add(batch)
    db.commit()

    response = client.post(
        "/api/dispatch/redispatch",
        json={"batch_ids": [batch.id], "category_code": "speaker"},
    )
    assert response.status_code == 202
    job_id = response.json()["job_id"]

    job = db.query(dispatch_api.DispatchRedispatchJob).filter_by(id=job_id).one()
    assert job.status == "done"
    assert job.failed_batches == 1
    item = db.query(dispatch_api.DispatchRedispatchJobItem).filter_by(job_id=job_id).one()
    assert item.status == "error"
    assert "无关联文件" in item.error_msg


def test_list_and_get_redispatch_jobs_are_scoped(client_and_db, monkeypatch):
    client, db = client_and_db
    _enable_sync_threads(monkeypatch)
    db.add_all([
        Category(code="headphone", name="耳机", sort_order=1),
        Category(code="speaker", name="音箱", sort_order=2),
    ])
    db.add(DispatchRule(
        category_code="speaker", platform="jd", field="item_name",
        match_type="contains", value="音箱", priority=1, is_active=1,
    ))
    db.commit()

    file_record, rows = _make_dispatch_file(db, filename="scoped-redispatch.xlsx", rows=[("手机配件", "无线音箱")])
    batch = _make_done_batch(db, file_record, category_code="headphone", raw_rows=rows)
    response = client.post(
        "/api/dispatch/redispatch",
        json={"batch_ids": [batch.id], "category_code": "speaker"},
    )
    assert response.status_code == 202
    job_id = response.json()["job_id"]

    _set_dispatch_user(client, category_permissions=["headphone"])
    assert client.get("/api/dispatch/redispatch/jobs").json()["total"] == 0
    assert client.get(f"/api/dispatch/redispatch/jobs/{job_id}").status_code == 403

    _set_dispatch_user(client)
    list_payload = client.get("/api/dispatch/redispatch/jobs").json()
    assert list_payload["total"] == 1
    assert list_payload["items"][0]["category_name"] == "音箱"
    assert list_payload["items"][0]["status"] == "done"

    detail_payload = client.get(f"/api/dispatch/redispatch/jobs/{job_id}").json()
    assert detail_payload["success_batches"] == 1
    assert len(detail_payload["items"]) == 1
    assert detail_payload["items"][0]["filename"] == "scoped-redispatch.xlsx"
