#!/usr/bin/env python3
"""
generate_dispatch_rules.py
从 平台元数据/RUNTO产品与电商类目对应关系.xlsx 解析分发规则，输出 dispatch_rules_import.sql

用法：
    python scripts/generate_dispatch_rules.py
"""
import os
import openpyxl
from collections import Counter

PRODUCT_MAP = {
    "智能投影":   "projector",
    "智能手表手环": "smartwatch",
    "智能眼镜":     "vrar",
    "XR和智能眼镜": "vrar",
    "移动电源":   "power_bank",
    "行车记录仪": "dashcam",
    "监控摄像头": "camera",
    "直播摄像头": "live_camera",
    "可视门铃":   "video_doorbell",
    "智能门锁":   "smart_lock",
    "路由器":     "router",
    "学习平板":   "edu_tablet",
    "智能平板":   "tablet",
    "词典笔":     "dict_pen",
    "移动智慧屏": "mobile_screen",
    "智能音箱":   "smart_speaker",
    "回音壁":     "soundbar",
    "无线蓝牙音箱": "bt_speaker",
    "耳机":       "headphone",
    "电子纸平板": "eink_tablet",
    "单词卡":     "word_card",
    "显示器":     "monitor",
    "笔记本":     "laptop",
}

PLATFORM_MAP = {
    "京东": "jd",
    "天猫": "tmall",
    "抖音": "douyin",
    "淘宝": "taobao",
}

FIELD_NAMES = {
    0: "category_lv0",
    1: "category_lv1",
    2: "category_lv2",
    3: "category_lv3",
    4: "category_lv4",
    5: "category_lv5",
}


def sql_str(s):
    """将 Python 值转为 SQL 字符串（带引号或 NULL）"""
    if s is None:
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


def parse_keywords(note: str) -> list[str]:
    """从备注中提取关键词列表，支持 、/ , 分隔"""
    if "按关键词搜索：" not in note:
        return []
    raw = note.split("按关键词搜索：", 1)[1].strip()
    return [k.strip() for k in raw.replace(",", "、").split("、") if k.strip()]


def parse_rules(excel_path: str) -> list[dict]:
    # 不用 read_only，以便正确展开合并单元格
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    # 展开合并单元格：将合并区域内每格填充为锚点格（左上角）的值
    for merge_range in list(ws.merged_cells.ranges):
        top_left_val = ws.cell(merge_range.min_row, merge_range.min_col).value
        ws.unmerge_cells(str(merge_range))
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                ws.cell(row, col).value = top_left_val

    all_rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    data_rows = all_rows[2:]   # 跳过两行表头

    rules: list[dict] = []
    seen: set[tuple] = set()  # (category_code, platform, field, value, item_name_keyword)

    cur_product: str | None = None
    cur_platform: str | None = None

    for row in data_rows:
        if not any(v is not None for v in row):
            continue

        product_raw = row[0]
        platform_raw = row[1]
        levels = [row[2], row[3], row[4], row[5], row[6], row[7]]  # lv0-lv5
        col_i = row[8]   # 宝贝名称/备注列

        # 产品向下填充
        if product_raw is not None:
            s = str(product_raw).strip()
            if s in PRODUCT_MAP:
                cur_product = s

        # 平台向下填充（仅当值为合法平台名时更新）
        if platform_raw is not None:
            s = str(platform_raw).strip()
            if s in PLATFORM_MAP:
                cur_platform = PLATFORM_MAP[s]

        if cur_product is None or cur_platform is None:
            continue

        category_code = PRODUCT_MAP.get(cur_product)
        if not category_code:
            continue

        # 找最深非空类目层级
        deepest_lv: int | None = None
        deepest_val: str | None = None
        for i in range(5, -1, -1):
            v = levels[i]
            if v is not None and str(v).strip():
                deepest_lv = i
                deepest_val = str(v).strip()
                break

        # 解析备注列
        note = str(col_i).strip() if col_i is not None else ""
        is_delete_replace = "删掉" in note
        keywords = parse_keywords(note)

        def add_rule(field, match_type, value, item_name_keyword, priority):
            key = (category_code, cur_platform, field, value, item_name_keyword)
            if key in seen:
                return
            seen.add(key)
            rules.append({
                "category_code": category_code,
                "platform": cur_platform,
                "field": field,
                "match_type": match_type,
                "value": value,
                "item_name_keyword": item_name_keyword,
                "priority": priority,
            })

        if is_delete_replace:
            # 用 item_name 关键词规则替代类目规则
            for kw in keywords:
                add_rule("item_name", "contains", kw, None, 80)

        elif deepest_lv is not None:
            field = FIELD_NAMES[deepest_lv]

            # 优先级：lv2+ 精确匹配=10，lv1+关键词=30，lv1=50，lv0+关键词=60，lv0=70
            if deepest_lv >= 2:
                base_priority = 10
            elif deepest_lv == 1:
                base_priority = 30 if keywords else 50
            else:  # lv0
                base_priority = 60 if keywords else 70

            if keywords:
                for kw in keywords:
                    add_rule(field, "equals", deepest_val, kw, base_priority)
            else:
                add_rule(field, "equals", deepest_val, None, base_priority)

    return rules


def generate_sql(rules: list[dict]) -> str:
    lines = [
        "-- ============================================================",
        "-- 分发规则导入 SQL（由 generate_dispatch_rules.py 自动生成）",
        "-- ============================================================",
        "SET NAMES utf8mb4;",
        "",
        "-- 1. 品类表更新",
        "INSERT IGNORE INTO categories (code, name) VALUES ('live_camera', '直播摄像头');",
        "UPDATE categories SET name='智能手表手环' WHERE code='smartwatch';",
        "UPDATE categories SET name='智能眼镜'    WHERE code='vrar';",
        "",
        f"-- 2. 分发规则（共 {len(rules)} 条）",
        "INSERT INTO dispatch_rules",
        "    (category_code, platform, field, match_type, value, item_name_keyword, priority, is_active)",
        "VALUES",
    ]

    value_rows = []
    for r in rules:
        value_rows.append(
            "  ({}, {}, {}, {}, {}, {}, {}, 1)".format(
                sql_str(r["category_code"]),
                sql_str(r["platform"]),
                sql_str(r["field"]),
                sql_str(r["match_type"]),
                sql_str(r["value"]),
                sql_str(r["item_name_keyword"]),
                r["priority"],
            )
        )
    lines.append(",\n".join(value_rows) + ";")
    return "\n".join(lines)


def main():
    base = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(base, "..", "平台元数据", "RUNTO产品与电商类目对应关系.xlsx")
    output_path = os.path.join(base, "dispatch_rules_import.sql")

    if not os.path.exists(excel_path):
        print(f"错误：找不到文件 {excel_path}")
        return

    rules = parse_rules(excel_path)
    print(f"解析完成，共 {len(rules)} 条规则\n")

    # 按品类+平台汇总
    counter = Counter((r["category_code"], r["platform"]) for r in rules)
    for (cat, plat), cnt in sorted(counter.items()):
        print(f"  {cat:20s} / {plat:8s}  {cnt:3d} 条")

    sql = generate_sql(rules)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(sql)
    print(f"\n已输出到 {output_path}")


if __name__ == "__main__":
    main()
