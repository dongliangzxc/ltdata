"""历史库管理 API
- POST   /api/historical/import           Excel 导入历史确认结果（upsert）
- GET    /api/historical/batches          查询所有批次名称列表
- GET    /api/historical/mappings         分页查询映射（platform / import_batch 筛选）
- DELETE /api/historical/mappings/batch   按批次批量删除（静态路由，必须在 /{id} 之前）
- DELETE /api/historical/mappings/{id}    删除单条映射
"""
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Optional
from uuid import UUID

import pandas as pd
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import func, or_, tuple_
from sqlalchemy.orm import Session

from app.core.auth_deps import get_current_user
from app.core.config import settings
from app.core.permissions import visible_category_codes
from app.models.database import get_db
from app.models.schemas import Category, HistoricalMapping, ModelRecord, User
from app.services.import_helper import save_tmp_file
from app.utils.time_utils import format_beijing_datetime
from app.utils.url_utils import extract_item_id

router = APIRouter(prefix="/api/historical", tags=["historical"])


class BatchDeleteIn(BaseModel):
    import_batch: str


class HistoricalConfirmIn(BaseModel):
    temp_file_id: str
    sheet_name: str
    mapping: dict[str, str]
    category_code: Optional[str] = None


class HistoricalPreviewIn(BaseModel):
    temp_file_id: str
    sheet_name: str
    mapping: dict[str, str]
    category_code: Optional[str] = None


def _visible_historical_category_codes(db: Session, current_user: User) -> set[str] | None:
    all_codes = [code for code, in db.query(Category.code).order_by(Category.sort_order, Category.name).all()]
    if not all_codes:
        return None
    return set(visible_category_codes(current_user, all_codes))


def _filter_historical_visible_categories(query, db: Session, current_user: User):
    visible_codes = _visible_historical_category_codes(db, current_user)
    if visible_codes is None:
        return query
    return query.filter(HistoricalMapping.category_code.in_(visible_codes))


def _ensure_historical_category_visible(db: Session, current_user: User, category_code: Optional[str]) -> None:
    if not category_code:
        return
    visible_codes = _visible_historical_category_codes(db, current_user)
    if visible_codes is not None and category_code not in visible_codes:
        raise HTTPException(status_code=403, detail="无权限访问该品类")


HISTORICAL_STANDARD_FIELDS = {
    "year": "年",
    "month_num": "月",
    "week": "周",
    "report_type": "报告类型",
    "channel": "渠道",
    "platform": "商场",
    "category_name_raw": "品类",
    "brand_raw": "品牌",
    "model_text": "型号",
    "category_code_raw": "品类码",
    "brand_code_raw": "品牌码",
    "model_code_raw": "型号码",
    "item_name": "标题",
    "sales_amount": "销额",
    "sales_qty": "销量",
    "price": "单价",
    "item_url": "网址",
}

HISTORICAL_REQUIRED_FIELDS = {"platform", "item_name", "year", "month_num"}

HISTORICAL_FIELD_ALIASES = {
    "year": {"年", "年度"},
    "month_num": {"月", "月度", "时间维度"},
    "week": {"周", "周度"},
    "report_type": {"报告类型"},
    "channel": {"渠道", "渠道类型"},
    "platform": {"商场", "平台"},
    "category_name_raw": {"品类", "类目", "产品品类"},
    "brand_raw": {"品牌"},
    "model_text": {"型号", "产品系列", "机型", "系列/机型", "品牌+系列", "品牌产品系列"},
    "category_code_raw": {"品类码"},
    "brand_code_raw": {"品牌码"},
    "model_code_raw": {"型号码"},
    "item_name": {"标题", "商品名称", "宝贝名称"},
    "sales_amount": {"销额", "销售额", "原始额"},
    "sales_qty": {"销量", "原始量"},
    "price": {"单价", "成交价"},
    "item_url": {"网址", "商品网址", "宝贝链接", "URL"},
}

HISTORICAL_ALIAS_TO_FIELD = {
    alias: field
    for field, aliases in HISTORICAL_FIELD_ALIASES.items()
    for alias in aliases
}


def _clean_value(value) -> Optional[str]:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.lower() == "nan":
        return None
    if text.endswith(".0"):
        try:
            return str(int(float(text)))
        except ValueError:
            return text
    return text


def _is_unusable_identity_value(value: Optional[str]) -> bool:
    text = _clean_value(value)
    if text is None:
        return True
    return text.strip().upper() in {"-", "—", "–", "UNKNOWN", "UNKNOWN BRAND", "未知"}


def _usable_identity_value(value: Optional[str]) -> Optional[str]:
    text = _clean_value(value)
    return None if _is_unusable_identity_value(text) else text


def _is_unknown_brand(value: Optional[str]) -> bool:
    return _is_unusable_identity_value(value)


def _clean_int(value) -> Optional[int]:
    text = _clean_value(value)
    if text is None:
        return None
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    if number != number.to_integral_value():
        return None
    return int(number)


def _parse_year(value) -> Optional[int]:
    year = _clean_int(value)
    if year is None:
        return None
    if 0 <= year < 100:
        return 2000 + year
    return year


def _parse_month(value) -> Optional[int]:
    text = _clean_value(value)
    if text is None:
        return None
    month = _clean_int(text)
    if month is not None:
        if 1 <= month <= 12:
            return month
        if text.isdigit() and len(text) == 6:
            month_part = month % 100
            return month_part if 1 <= month_part <= 12 else month
        return month

    normalized = text.replace("年", ".").replace("月", "").replace("-", ".").replace("/", ".")
    if "." not in normalized:
        return None
    year_text, month_text = normalized.split(".", 1)
    if len(year_text.strip()) != 4:
        return None
    return _clean_int(month_text)


def _parse_time_dimension(value) -> tuple[Optional[int], Optional[int]]:
    text = _clean_value(value)
    if text is None:
        return None, None
    normalized = text.replace("年", ".").replace("月", "").replace("-", ".").replace("/", ".")
    if "." not in normalized:
        return None, None
    year_text, month_text = normalized.split(".", 1)
    year = _parse_year(year_text)
    month = _clean_int(month_text)
    return year, month


def _clean_decimal(value) -> Optional[Decimal]:
    text = _clean_value(value)
    if text is None:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _normalize_platform(value: Optional[str]) -> Optional[str]:
    text = _clean_value(value)
    return text.lower() if text else None


def _normalize_item_name(value: Optional[str]) -> Optional[str]:
    text = _clean_value(value)
    return " ".join(text.upper().split()) if text else None


def _get(row: dict, key: str):
    return row.get(key)


def _row_value(row: dict, mapping: dict[str, str], field: str):
    source = mapping.get(field)
    if source is None:
        return None
    return row.get(source)


def _build_mapping(columns: list[str]) -> dict[str, str]:
    mapping = {}
    used = set()
    for col in columns:
        field = HISTORICAL_ALIAS_TO_FIELD.get(col)
        if field and field not in used:
            mapping[field] = col
            used.add(field)
        elif field == "item_url" and col == "网址":
            mapping[field] = col
    return mapping


def _read_sheet_preview(
    xls: pd.ExcelFile,
    sheet_name: str,
    nrows: Optional[int] = None,
    usecols: Optional[list[str]] = None,
) -> pd.DataFrame:
    df = pd.read_excel(xls, sheet_name=sheet_name, nrows=nrows, usecols=usecols)
    df.columns = [str(col).strip() for col in df.columns]
    return df


def _iter_sheet_rows_streaming(
    path: Path,
    sheet_name: str,
    *,
    columns: list[str],
    usecols: Optional[list[str]] = None,
):
    selected = set(usecols or columns)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise HTTPException(400, "选择的 sheet 不存在")
        sheet = workbook[sheet_name]
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return
        header_values = [str(col).strip() if col is not None else "" for col in header]
        selected_indexes = [(index, col) for index, col in enumerate(header_values) if col in selected]
        for row in rows_iter:
            yield {col: row[index] if index < len(row) else None for index, col in selected_indexes}
    finally:
        workbook.close()


def _read_sheet_streaming(
    path: Path,
    sheet_name: str,
    *,
    columns: list[str],
    nrows: Optional[int] = None,
    usecols: Optional[list[str]] = None,
) -> pd.DataFrame:
    rows = []
    for row in _iter_sheet_rows_streaming(path, sheet_name, columns=columns, usecols=usecols):
        rows.append(row)
        if nrows is not None and len(rows) >= nrows:
            break
    return pd.DataFrame(rows, columns=usecols or columns)


def _sheet_total_rows_streaming(path: Path, sheet_name: str) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise HTTPException(400, "选择的 sheet 不存在")
        sheet = workbook[sheet_name]
        return max((sheet.max_row or 1) - 1, 0)
    finally:
        workbook.close()


def _mapped_columns(mapping: dict[str, str], columns: list[str]) -> list[str]:
    column_set = set(columns)
    return list(dict.fromkeys(source for source in mapping.values() if source in column_set))


def _score_sheet(columns: list[str]) -> int:
    mapping = _build_mapping(columns)
    score = len(mapping)
    score += sum(5 for field in HISTORICAL_REQUIRED_FIELDS if field in mapping)
    if "item_url" in mapping:
        score += 3
    if "sales_qty" in mapping or "sales_amount" in mapping:
        score += 2
    return score


def _detect_sheet_and_mapping(xls: pd.ExcelFile) -> tuple[str, list[str], dict[str, str]]:
    candidates = []
    for sheet_name in xls.sheet_names:
        try:
            df = _read_sheet_preview(xls, sheet_name, nrows=0)
        except Exception:
            continue
        columns = [str(col).strip() for col in df.columns]
        candidates.append((_score_sheet(columns), sheet_name, columns, _build_mapping(columns)))
    if not candidates:
        raise HTTPException(400, "Excel 文件没有可读取的工作表")
    rawdata = [item for item in candidates if item[1] == "rawdata"]
    if rawdata:
        _, sheet_name, columns, mapping = rawdata[0]
        return sheet_name, columns, mapping
    candidates.sort(key=lambda item: item[0], reverse=True)
    _, sheet_name, columns, mapping = candidates[0]
    return sheet_name, columns, mapping


LEGACY_FILENAME_CATEGORY_ALIASES = {
    "door_lock": ("门锁",),
    "smart_lock": ("门锁",),
}


def _infer_category_code(filename: str, db: Session) -> Optional[str]:
    categories = db.query(Category).all()
    for category in categories:
        aliases = LEGACY_FILENAME_CATEGORY_ALIASES.get(category.code, ())
        if any(alias in filename for alias in aliases):
            return category.code
        if category.name and category.name in filename:
            return category.code
        if category.code and category.code in filename:
            return category.code
    return None


def _standardize_historical_row(row: dict, mapping: dict[str, str], category_code: Optional[str]) -> dict:
    year_value = _row_value(row, mapping, "year")
    month_value = _row_value(row, mapping, "month_num")
    year = _parse_year(year_value)
    month_num = _parse_month(month_value)
    if (year is None or month_num is None) and mapping.get("month_num") in {"时间维度", "月度"}:
        dimension_year, dimension_month = _parse_time_dimension(month_value)
        year = year if year is not None else dimension_year
        month_num = month_num if month_num is not None else dimension_month
    return {
        "年": year,
        "月": month_num,
        "周": _clean_value(_row_value(row, mapping, "week")),
        "报告类型": _clean_value(_row_value(row, mapping, "report_type")),
        "渠道": _clean_value(_row_value(row, mapping, "channel")),
        "商场": _clean_value(_row_value(row, mapping, "platform")),
        "品类": _clean_value(_row_value(row, mapping, "category_name_raw")),
        "品牌": _clean_value(_row_value(row, mapping, "brand_raw")),
        "型号": _clean_value(_row_value(row, mapping, "model_text")),
        "品类码": _clean_value(_row_value(row, mapping, "category_code_raw")) or category_code,
        "品牌码": _clean_value(_row_value(row, mapping, "brand_code_raw")),
        "型号码": _clean_value(_row_value(row, mapping, "model_code_raw")),
        "标题": _clean_value(_row_value(row, mapping, "item_name")),
        "销额": _clean_value(_row_value(row, mapping, "sales_amount")),
        "销量": _clean_value(_row_value(row, mapping, "sales_qty")),
        "单价": _clean_value(_row_value(row, mapping, "price")),
        "网址": _clean_value(_row_value(row, mapping, "item_url")),
    }


def _standardize_historical_df(df: pd.DataFrame, mapping: dict[str, str], category_code: Optional[str]) -> pd.DataFrame:
    rows = []
    for _, raw_row in df.iterrows():
        rows.append(_standardize_historical_row(raw_row.to_dict(), mapping, category_code))
    return pd.DataFrame(rows)


def _preview_rows(df: pd.DataFrame, limit: int = 20) -> list[dict]:
    rows = []
    for _, row in df.head(limit).iterrows():
        rows.append(_json_safe_row(row.to_dict()))
    return rows


def _normalize_mapping_for_columns(mapping: dict[str, str], columns: list[str]) -> dict[str, str]:
    column_set = set(columns)
    normalized = {field: source for field, source in mapping.items() if source in column_set}
    defaults = _build_mapping(columns)
    for field, source in defaults.items():
        normalized.setdefault(field, source)
    return normalized


def _mapping_issues(mapping: dict[str, str], category_code: Optional[str], columns: Optional[list[str]] = None) -> list[str]:
    issues = []
    column_set = set(columns) if columns is not None else None
    for field, source in mapping.items():
        if column_set is not None and source not in column_set:
            label = HISTORICAL_STANDARD_FIELDS.get(field, field)
            issues.append(f"字段映射不存在：{label} -> {source}")
    for field in sorted(HISTORICAL_REQUIRED_FIELDS):
        if field == "year" and mapping.get("month_num") == "时间维度":
            continue
        if field not in mapping:
            issues.append(f"缺少必填字段映射：{HISTORICAL_STANDARD_FIELDS[field]}")
    if not category_code and "category_code_raw" not in mapping and "category_name_raw" not in mapping:
        issues.append("未识别品类，确认导入前需要选择品类")
    return issues


def _row_required_errors(row: dict) -> list[str]:
    errors = []
    platform = _normalize_platform(_get(row, "商场")) or _normalize_platform(_get(row, "渠道"))
    item_name = _clean_value(_get(row, "标题"))
    year = _parse_year(_get(row, "年"))
    month_num = _parse_month(_get(row, "月"))
    if month_num is None:
        _dimension_year, dimension_month = _parse_time_dimension(_get(row, "月"))
        month_num = dimension_month
    if not platform:
        errors.append("商场/渠道不能为空")
    if not item_name:
        errors.append("标题不能为空")
    if year is None:
        errors.append("年不能为空")
    if month_num is None:
        errors.append("月不能为空")
    elif month_num < 1 or month_num > 12:
        errors.append("月必须为 1-12")
    return errors


def _effective_brand_code(brand_code_raw: Optional[str], brand_raw: Optional[str]) -> Optional[str]:
    if not _is_unknown_brand(brand_code_raw):
        return brand_code_raw
    if not _is_unknown_brand(brand_raw):
        return brand_raw
    return None


def _preview_stats(db: Session, standardized_df: pd.DataFrame) -> dict:
    return _preview_stats_for_rows(db, (raw_row.to_dict() for _, raw_row in standardized_df.iterrows()))


def _preview_stats_for_rows(db: Session, rows: Iterable[dict]) -> dict:
    materialized_rows = list(rows)
    model_code_values = {_clean_value(_get(row, "型号码")) for row in materialized_rows}
    model_text_values = {_clean_value(_get(row, "型号")) for row in materialized_rows}
    models_by_code, models_by_code_unbranded, models_by_name = _load_preview_models(db, model_code_values, model_text_values)
    ambiguous_model_codes = _ambiguous_model_codes_from_rows(materialized_rows, models_by_code_unbranded)
    return _calculate_preview_stats_for_rows(
        materialized_rows,
        models_by_code=models_by_code,
        models_by_code_unbranded=models_by_code_unbranded,
        models_by_name=models_by_name,
        ambiguous_model_codes=ambiguous_model_codes,
    )


def _load_preview_models(db: Session, model_code_values: set[str | None], model_text_values: set[str | None]):
    model_code_values.discard(None)
    model_text_values.discard(None)
    all_codes = model_code_values | model_text_values

    models_by_code = {}
    models_by_code_unbranded = {}
    if all_codes:
        for model in db.query(ModelRecord).filter(ModelRecord.model_code.in_(all_codes)).all():
            models_by_code[(model.brand_code, model.model_code)] = model
            models_by_code_unbranded.setdefault(model.model_code, []).append(model)

    models_by_name = {}
    if model_text_values:
        for model in db.query(ModelRecord).filter(ModelRecord.model_name.in_(model_text_values)).all():
            models_by_name.setdefault(model.model_name, []).append(model)

    return models_by_code, models_by_code_unbranded, models_by_name


def _calculate_preview_stats_for_rows(
    rows: Iterable[dict],
    *,
    models_by_code: dict[tuple[str, str], ModelRecord],
    models_by_code_unbranded: dict[str, list[ModelRecord]],
    models_by_name: dict[str, list[ModelRecord]],
    ambiguous_model_codes: set[str],
) -> dict:
    total_rows = 0
    missing_required_rows = 0
    missing_model_rows = 0
    auto_create_keys: set[tuple[str, str]] = set()

    for row in rows:
        total_rows += 1
        if _row_required_errors(row):
            missing_required_rows += 1
            continue
        model_code_raw = _clean_value(_get(row, "型号码"))
        model_text = _clean_value(_get(row, "型号"))
        brand_raw = _clean_value(_get(row, "品牌"))
        brand_code_raw = _clean_value(_get(row, "品牌码"))
        effective_brand_code = _effective_brand_code(brand_code_raw, brand_raw)
        if not (model_code_raw or model_text):
            continue
        model, reason = _resolve_model(
            model_code_raw=model_code_raw,
            model_text=model_text,
            brand_code_raw=effective_brand_code,
            models_by_code=models_by_code,
            models_by_code_unbranded=models_by_code_unbranded,
            models_by_name=models_by_name,
            ambiguous_model_codes=ambiguous_model_codes,
        )
        if model is None:
            missing_model_rows += 1
            if reason and ("匹配到多个型号" in reason or "匹配到多个品牌" in reason):
                continue
            values = _model_identity_values(
                brand_code_raw=effective_brand_code,
                brand_raw=brand_raw,
                model_code_raw=model_code_raw,
                model_text=model_text,
                category_code=_clean_value(_get(row, "品类码")),
            )
            if values is not None:
                auto_create_keys.add((values["brand_code"], values["model_code"]))

    return {
        "total_rows": total_rows,
        "importable_rows": total_rows - missing_required_rows,
        "missing_required_rows": missing_required_rows,
        "missing_model_rows": missing_model_rows,
        "auto_create_model_count": len(auto_create_keys),
    }


def _resolve_model(
    *,
    model_code_raw: Optional[str],
    model_text: Optional[str],
    brand_code_raw: Optional[str],
    models_by_code: dict[tuple[str, str], ModelRecord],
    models_by_code_unbranded: dict[str, list[ModelRecord]],
    models_by_name: dict[str, list[ModelRecord]],
    ambiguous_model_codes: Optional[set[str]] = None,
):
    ambiguous_model_codes = ambiguous_model_codes or set()
    if model_code_raw:
        if not brand_code_raw and model_code_raw in ambiguous_model_codes:
            return None, f"型号码「{model_code_raw}」匹配到多个品牌，请填写品牌码"
        if brand_code_raw:
            model = models_by_code.get((brand_code_raw, model_code_raw))
            if model:
                return model, None
        else:
            models_by_code_list = models_by_code_unbranded.get(model_code_raw, [])
            if len(models_by_code_list) == 1:
                return models_by_code_list[0], None
            if len(models_by_code_list) > 1:
                return None, f"型号码「{model_code_raw}」匹配到多个品牌，请填写品牌码"
        if not model_text:
            return None, f"型号码「{model_code_raw}」在型号库中不存在"

    if not model_text:
        return None, None

    if not brand_code_raw and model_text in ambiguous_model_codes:
        return None, f"型号码「{model_text}」匹配到多个品牌，请填写品牌码"

    if brand_code_raw:
        model_by_code = models_by_code.get((brand_code_raw, model_text))
        if model_by_code:
            return model_by_code, None
    else:
        models_by_code_list = models_by_code_unbranded.get(model_text, [])
        if len(models_by_code_list) == 1:
            return models_by_code_list[0], None
        if len(models_by_code_list) > 1:
            return None, f"型号码「{model_text}」匹配到多个品牌，请填写品牌码"

    models_by_name_list = models_by_name.get(model_text, [])
    if len(models_by_name_list) == 1:
        if not brand_code_raw or models_by_name_list[0].brand_code == brand_code_raw:
            return models_by_name_list[0], None
    if len(models_by_name_list) > 1:
        if brand_code_raw:
            matched = [m for m in models_by_name_list if m.brand_code == brand_code_raw]
            if len(matched) == 1:
                return matched[0], None
        return None, f"型号名称「{model_text}」匹配到多个型号，请填写型号码"

    return None, f"型号「{model_text}」在型号库中不存在"


def _match_key_type(item_id: Optional[str], item_url: Optional[str], item_name_norm: str) -> str:
    if item_id:
        return "item_id"
    if item_url:
        return "item_url"
    return "item_name"


def _period_filter(query, *, year: int, month_num: int, week: Optional[str]):
    query = query.filter(
        HistoricalMapping.year == year,
        HistoricalMapping.month_num == month_num,
    )
    if week:
        query = query.filter(HistoricalMapping.week == week)
    else:
        query = query.filter(HistoricalMapping.week.is_(None))
    return query


def _history_identity_key(
    *,
    platform: str,
    item_id: Optional[str],
    item_url: Optional[str],
    item_name_norm: str,
    year: int,
    month_num: int,
    week: Optional[str],
) -> tuple[str, str, int, int, Optional[str], str]:
    if item_id:
        return platform, "item_id", year, month_num, week, item_id
    if item_url:
        return platform, "item_url", year, month_num, week, item_url
    return platform, "item_name", year, month_num, week, item_name_norm


def _chunks(values: set[tuple], size: int = 1000):
    batch = []
    for value in values:
        batch.append(value)
        if len(batch) >= size:
            yield batch
            batch = []
    if batch:
        yield batch


def _preload_existing_history(db: Session, keys: set[tuple[str, str, int, int, Optional[str], str]]) -> dict[tuple[str, str, int, int, Optional[str], str], HistoricalMapping]:
    if not keys:
        return {}

    item_keys = {(platform, year, month_num, value) for platform, key_type, year, month_num, week, value in keys if key_type == "item_id"}
    url_keys = {(platform, year, month_num, value) for platform, key_type, year, month_num, week, value in keys if key_type == "item_url"}
    name_keys = {(platform, year, month_num, value) for platform, key_type, year, month_num, week, value in keys if key_type == "item_name"}

    existing = {}

    def add_rows(rows, key_type: str, value_attr: str):
        for row in rows:
            value = getattr(row, value_attr)
            key = (row.platform, key_type, row.year, row.month_num, row.week, value)
            existing[key] = row

    for batch in _chunks(item_keys):
        rows = db.query(HistoricalMapping).filter(
            tuple_(HistoricalMapping.platform, HistoricalMapping.year, HistoricalMapping.month_num, HistoricalMapping.item_id).in_(batch)
        ).all()
        add_rows(rows, "item_id", "item_id")
    for batch in _chunks(url_keys):
        rows = db.query(HistoricalMapping).filter(
            tuple_(HistoricalMapping.platform, HistoricalMapping.year, HistoricalMapping.month_num, HistoricalMapping.item_url).in_(batch)
        ).all()
        add_rows(rows, "item_url", "item_url")
    for batch in _chunks(name_keys):
        rows = db.query(HistoricalMapping).filter(
            tuple_(HistoricalMapping.platform, HistoricalMapping.year, HistoricalMapping.month_num, HistoricalMapping.item_name_norm).in_(batch)
        ).all()
        add_rows(rows, "item_name", "item_name_norm")
    return existing


def _json_safe_row(row: dict) -> dict:
    result = {}
    for key, value in row.items():
        cleaned = _clean_value(value)
        result[key] = cleaned
    return result


def _sum_optional_int(left: Optional[int], right: Optional[int]) -> Optional[int]:
    if left is None:
        return right
    if right is None:
        return left
    return left + right


def _sum_optional_decimal(left: Optional[Decimal], right: Optional[Decimal]) -> Optional[Decimal]:
    if left is None:
        return right
    if right is None:
        return left
    return left + right


def _merge_duplicate_history_values(current: dict, incoming: dict, row_nums: list[int]) -> None:
    sales_qty = _sum_optional_int(current.get("sales_qty"), incoming.get("sales_qty"))
    sales_amount = _sum_optional_decimal(current.get("sales_amount"), incoming.get("sales_amount"))
    current.update(incoming)
    current["sales_qty"] = sales_qty
    current["sales_amount"] = sales_amount
    raw_payload = dict(current.get("raw_payload") or {})
    raw_payload["_merge_note"] = f"同一导入批次内发现 {len(row_nums)} 条相同历史键记录，销量和销额已累加"
    raw_payload["_merged_rows"] = row_nums
    current["raw_payload"] = raw_payload


def _preload_categories(db: Session) -> tuple[dict[str, str], dict[str, str]]:
    by_code = {}
    by_name = {}
    for category in db.query(Category).all():
        by_code[category.code] = category.code
        by_name[category.name] = category.code
    return by_code, by_name


def _resolve_category_code(
    *,
    category_code_raw: Optional[str],
    category_name_raw: Optional[str],
    categories_by_code: dict[str, str],
    categories_by_name: dict[str, str],
) -> Optional[str]:
    if category_code_raw and category_code_raw in categories_by_code:
        return categories_by_code[category_code_raw]
    if category_name_raw and category_name_raw in categories_by_name:
        return categories_by_name[category_name_raw]
    return None


def _pending_model_identity(category_code: Optional[str]) -> tuple[str, str]:
    suffix = category_code or "unknown"
    return f"待补型号-{suffix}", "待补型号"


def _model_identity_values(
    *,
    brand_code_raw: Optional[str],
    brand_raw: Optional[str],
    model_code_raw: Optional[str],
    model_text: Optional[str],
    category_code: Optional[str],
    allow_pending_model: bool = False,
) -> Optional[dict]:
    brand_code = _effective_brand_code(brand_code_raw, brand_raw)
    if _is_unknown_brand(brand_code):
        return None
    model_code = _usable_identity_value(model_code_raw)
    model_name = _usable_identity_value(model_text)
    if not model_code and not model_name:
        if not allow_pending_model:
            return None
        model_code, model_name = _pending_model_identity(category_code)
    else:
        model_code = model_code or model_name
        model_name = model_name or model_code
    return {
        "brand_code": brand_code,
        "model_code": model_code,
        "model_name": model_name,
        "brand_name": _usable_identity_value(brand_raw) or brand_code,
        "category_code": category_code,
    }


def _get_or_create_model(
    db: Session,
    *,
    model_code_raw: Optional[str],
    model_text: Optional[str],
    brand_code_raw: Optional[str],
    brand_raw: Optional[str],
    category_code: Optional[str],
    models_by_code: dict[tuple[str, str], ModelRecord],
    models_by_code_unbranded: dict[str, list[ModelRecord]],
    models_by_name: dict[str, list[ModelRecord]],
    ambiguous_model_codes: Optional[set[str]] = None,
) -> tuple[Optional[ModelRecord], Optional[str]]:
    model, reason = _resolve_model(
        model_code_raw=model_code_raw,
        model_text=model_text,
        brand_code_raw=brand_code_raw,
        models_by_code=models_by_code,
        models_by_code_unbranded=models_by_code_unbranded,
        models_by_name=models_by_name,
        ambiguous_model_codes=ambiguous_model_codes,
    )
    if model:
        return model, None
    if not reason:
        return None, None
    if "匹配到多个型号" in reason or "匹配到多个品牌" in reason:
        return None, reason

    values = _model_identity_values(
        brand_code_raw=brand_code_raw,
        brand_raw=brand_raw,
        model_code_raw=model_code_raw,
        model_text=model_text,
        category_code=category_code,
    )
    if values is None:
        return None, None

    existing = db.query(ModelRecord).filter(
        ModelRecord.brand_code == values["brand_code"],
        ModelRecord.model_code == values["model_code"],
    ).first()
    if existing:
        models_by_code[(existing.brand_code, existing.model_code)] = existing
        if existing not in models_by_code_unbranded.setdefault(existing.model_code, []):
            models_by_code_unbranded[existing.model_code].append(existing)
        if existing not in models_by_name.setdefault(existing.model_name, []):
            models_by_name[existing.model_name].append(existing)
        return existing, None

    new_model = ModelRecord(**values)
    db.add(new_model)
    db.flush()
    models_by_code[(new_model.brand_code, new_model.model_code)] = new_model
    models_by_code_unbranded.setdefault(new_model.model_code, []).append(new_model)
    models_by_name.setdefault(new_model.model_name, []).append(new_model)
    return new_model, None


def _find_tmp_file(temp_file_id: str) -> Path:
    try:
        safe_id = str(UUID(temp_file_id))
    except ValueError:
        raise HTTPException(400, "无效的临时文件 ID")
    tmp_dir = Path(settings.UPLOAD_DIR) / "tmp"
    matches = list(tmp_dir.glob(f"{safe_id}_*"))
    if not matches:
        raise HTTPException(404, "临时文件不存在或已过期，请重新上传")
    return matches[0]


def _original_filename_from_tmp(path: Path, temp_file_id: str) -> str:
    prefix = f"{temp_file_id}_"
    if path.name.startswith(prefix):
        return path.name[len(prefix):]
    return path.name


def _validate_excel_file(file: UploadFile):
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "仅支持 Excel 文件（.xlsx/.xls）")


def _historical_preview_response(
    *,
    temp_file_id: str,
    filename: str,
    xls: pd.ExcelFile,
    path: Path,
    sheet_name: str,
    columns: list[str],
    mapping: dict[str, str],
    category_code: Optional[str],
    db: Session,
) -> dict:
    raw_issues = _mapping_issues(mapping, category_code, columns)
    normalized_mapping = _normalize_mapping_for_columns(mapping, columns)
    normalized_issues = _mapping_issues(normalized_mapping, category_code, columns)
    issues = list(dict.fromkeys(raw_issues + normalized_issues))
    usecols = _mapped_columns(normalized_mapping, columns)
    if path.suffix.lower() == ".xls":
        df_preview = _read_sheet_preview(xls, sheet_name, nrows=20, usecols=usecols)
        standardized_preview = _standardize_historical_df(df_preview, normalized_mapping, category_code)
        full_df = _standardize_historical_df(_read_sheet_preview(xls, sheet_name, usecols=usecols), normalized_mapping, category_code)
        stats = _preview_stats(db, full_df)
        total_rows = stats["total_rows"]
        preview_rows = _preview_rows(standardized_preview)
    else:
        preview_rows = []
        model_code_values = set()
        model_text_values = set()
        for raw_row in _iter_sheet_rows_streaming(path, sheet_name, columns=columns, usecols=usecols):
            standardized = _standardize_historical_row(raw_row, normalized_mapping, category_code)
            model_code_values.add(_clean_value(_get(standardized, "型号码")))
            model_text_values.add(_clean_value(_get(standardized, "型号")))
            if len(preview_rows) < 20:
                preview_rows.append(_json_safe_row(standardized))
        models_by_code, models_by_code_unbranded, models_by_name = _load_preview_models(db, model_code_values, model_text_values)
        ambiguous_model_codes = _ambiguous_model_codes_from_rows(
            (
                _standardize_historical_row(raw_row, normalized_mapping, category_code)
                for raw_row in _iter_sheet_rows_streaming(path, sheet_name, columns=columns, usecols=usecols)
            ),
            models_by_code_unbranded,
        )
        stats = _calculate_preview_stats_for_rows(
            (
                _standardize_historical_row(raw_row, normalized_mapping, category_code)
                for raw_row in _iter_sheet_rows_streaming(path, sheet_name, columns=columns, usecols=usecols)
            ),
            models_by_code=models_by_code,
            models_by_code_unbranded=models_by_code_unbranded,
            models_by_name=models_by_name,
            ambiguous_model_codes=ambiguous_model_codes,
        )
        total_rows = stats["total_rows"]
    return {
        "temp_file_id": temp_file_id,
        "filename": filename,
        "sheets": xls.sheet_names,
        "sheet_name": sheet_name,
        "columns": columns,
        "standard_fields": HISTORICAL_STANDARD_FIELDS,
        "mapping": normalized_mapping,
        "category_code": category_code,
        "issues": issues,
        "total_rows": total_rows,
        "stats": stats,
        "preview": preview_rows,
    }


def _preload_models(db: Session, df: pd.DataFrame):
    model_code_values = {_clean_value(value) for value in df.get("型号码", [])}
    model_text_values = {_clean_value(value) for value in df.get("型号", [])}
    model_code_values.discard(None)
    model_text_values.discard(None)
    all_codes = model_code_values | model_text_values

    models_by_code = {}
    models_by_code_unbranded = {}
    if all_codes:
        for model in db.query(ModelRecord).filter(ModelRecord.model_code.in_(all_codes)).all():
            models_by_code[(model.brand_code, model.model_code)] = model
            models_by_code_unbranded.setdefault(model.model_code, []).append(model)

    models_by_name = {}
    if model_text_values:
        for model in db.query(ModelRecord).filter(ModelRecord.model_name.in_(model_text_values)).all():
            models_by_name.setdefault(model.model_name, []).append(model)

    return models_by_code, models_by_code_unbranded, models_by_name


def _ambiguous_model_codes_from_rows(rows: Iterable[dict], models_by_code_unbranded: dict[str, list[ModelRecord]]) -> set[str]:
    brands_by_code = {
        model_code: {model.brand_code for model in models}
        for model_code, models in models_by_code_unbranded.items()
    }
    for row in rows:
        brand_code_raw = _clean_value(_get(row, "品牌码"))
        brand_raw = _clean_value(_get(row, "品牌"))
        brand_code = _effective_brand_code(brand_code_raw, brand_raw)
        if _is_unknown_brand(brand_code):
            continue
        for model_code in {_clean_value(_get(row, "型号码")), _clean_value(_get(row, "型号"))}:
            if model_code:
                brands_by_code.setdefault(model_code, set()).add(brand_code)
    return {model_code for model_code, brands in brands_by_code.items() if len(brands) > 1}


def _ambiguous_model_codes_from_batch(df: pd.DataFrame, models_by_code_unbranded: dict[str, list[ModelRecord]]) -> set[str]:
    return _ambiguous_model_codes_from_rows((raw_row.to_dict() for _, raw_row in df.iterrows()), models_by_code_unbranded)


def _import_historical_dataframe(db: Session, df: pd.DataFrame, import_batch: str) -> dict:
    if df.empty:
        df = pd.DataFrame([{col: None for col in df.columns}])
    rows = (raw_row.to_dict() for _, raw_row in df.iterrows())
    models_by_code, models_by_code_unbranded, models_by_name = _preload_models(db, df)
    ambiguous_model_codes = _ambiguous_model_codes_from_batch(df, models_by_code_unbranded)
    return _import_historical_rows(
        db,
        rows,
        import_batch,
        models_by_code=models_by_code,
        models_by_code_unbranded=models_by_code_unbranded,
        models_by_name=models_by_name,
        ambiguous_model_codes=ambiguous_model_codes,
    )


def _model_lookup_values_from_rows(rows: Iterable[dict]) -> tuple[set[str | None], set[str | None]]:
    model_code_values = set()
    model_text_values = set()
    for row in rows:
        model_code_values.add(_clean_value(_get(row, "型号码")))
        model_text_values.add(_clean_value(_get(row, "型号")))
    return model_code_values, model_text_values


def _historical_values_from_existing(existing: HistoricalMapping) -> dict:
    return {
        "import_batch": existing.import_batch,
        "platform": existing.platform,
        "item_id": existing.item_id,
        "item_url": existing.item_url,
        "item_name": existing.item_name,
        "item_name_norm": existing.item_name_norm,
        "year": existing.year,
        "month_num": existing.month_num,
        "week": existing.week,
        "month": existing.month,
        "report_type": existing.report_type,
        "channel": existing.channel,
        "category_name_raw": existing.category_name_raw,
        "category_code_raw": existing.category_code_raw,
        "brand_raw": existing.brand_raw,
        "brand_code_raw": existing.brand_code_raw,
        "model_text": existing.model_text,
        "model_code_raw": existing.model_code_raw,
        "model_id": existing.model_id,
        "model_code": existing.model_code,
        "category_code": existing.category_code,
        "sales_amount": existing.sales_amount,
        "sales_qty": existing.sales_qty,
        "price": existing.price,
        "match_key_type": existing.match_key_type,
        "raw_payload": existing.raw_payload,
        "updated_at": existing.updated_at,
    }


def _apply_historical_batch(
    db: Session,
    merged_rows: dict[tuple[str, str, int, int, Optional[str], str], dict],
    merged_row_nums: dict[tuple[str, str, int, int, Optional[str], str], list[int]],
    seen_current_keys: set[tuple[str, str, int, int, Optional[str], str]],
    seen_row_nums: dict[tuple[str, str, int, int, Optional[str], str], list[int]],
) -> tuple[int, int]:
    created = 0
    updated = 0
    touched_rows = []
    existing_by_key = _preload_existing_history(db, set(merged_rows))
    for history_key, values in merged_rows.items():
        row_nums = merged_row_nums[history_key]
        existing = existing_by_key.get(history_key)
        if existing:
            touched_rows.append(existing)
            if history_key in seen_current_keys:
                all_row_nums = seen_row_nums.setdefault(history_key, [])
                all_row_nums.extend(row_nums)
                merged_values = _historical_values_from_existing(existing)
                _merge_duplicate_history_values(merged_values, values, all_row_nums)
                for key, value in merged_values.items():
                    setattr(existing, key, value)
            else:
                for key, value in values.items():
                    setattr(existing, key, value)
                updated += 1
        else:
            new_row = HistoricalMapping(**values)
            db.add(new_row)
            touched_rows.append(new_row)
            created += 1
        if history_key not in seen_current_keys:
            seen_current_keys.add(history_key)
            seen_row_nums[history_key] = list(row_nums)
        elif history_key not in seen_row_nums:
            seen_row_nums[history_key] = list(row_nums)
    db.commit()
    for row in touched_rows:
        if row in db:
            db.expunge(row)
    return created, updated


def _import_historical_rows(
    db: Session,
    rows: Iterable[dict],
    import_batch: str,
    *,
    models_by_code: dict[tuple[str, str], ModelRecord],
    models_by_code_unbranded: dict[str, list[ModelRecord]],
    models_by_name: dict[str, list[ModelRecord]],
    ambiguous_model_codes: set[str],
) -> dict:
    success = 0
    created = 0
    updated = 0
    errors = []
    categories_by_code, categories_by_name = _preload_categories(db)
    merged_rows: dict[tuple[str, str, int, int, Optional[str], str], dict] = {}
    merged_row_nums: dict[tuple[str, str, int, int, Optional[str], str], list[int]] = {}
    seen_current_keys: set[tuple[str, str, int, int, Optional[str], str]] = set()
    seen_row_nums: dict[tuple[str, str, int, int, Optional[str], str], list[int]] = {}

    def flush_batch() -> None:
        nonlocal created, updated, merged_rows, merged_row_nums
        if not merged_rows:
            return
        batch_created, batch_updated = _apply_historical_batch(
            db,
            merged_rows,
            merged_row_nums,
            seen_current_keys,
            seen_row_nums,
        )
        created += batch_created
        updated += batch_updated
        merged_rows = {}
        merged_row_nums = {}

    for idx, row in enumerate(rows):
        row_num = idx + 2
        platform = _normalize_platform(_get(row, "商场")) or _normalize_platform(_get(row, "渠道"))
        item_name = _clean_value(_get(row, "标题"))
        model_text = _clean_value(_get(row, "型号"))
        year = _parse_year(_get(row, "年"))
        month_num = _parse_month(_get(row, "月"))

        missing = []
        if not platform:
            missing.append("商场/渠道不能为空")
        if not item_name:
            missing.append("标题不能为空")
        if year is None:
            missing.append("年不能为空")
        if month_num is None:
            missing.append("月不能为空")
        if missing:
            errors.append({"row": row_num, "reason": "；".join(missing)})
            continue
        if month_num < 1 or month_num > 12:
            errors.append({"row": row_num, "reason": "月必须为 1-12"})
            continue

        item_url = _clean_value(_get(row, "网址"))
        parsed = extract_item_id(item_url)
        item_id = None
        if parsed:
            parsed_platform, item_id = parsed
            platform = parsed_platform

        brand_raw = _clean_value(_get(row, "品牌"))
        brand_code_raw = _clean_value(_get(row, "品牌码"))
        effective_brand_code = _effective_brand_code(brand_code_raw, brand_raw)
        model_code_raw = _clean_value(_get(row, "型号码"))
        category_name_raw = _clean_value(_get(row, "品类"))
        category_code_raw = _clean_value(_get(row, "品类码"))
        resolved_category_code = _resolve_category_code(
            category_code_raw=category_code_raw,
            category_name_raw=category_name_raw,
            categories_by_code=categories_by_code,
            categories_by_name=categories_by_name,
        )
        model, reason = _get_or_create_model(
            db,
            model_code_raw=model_code_raw,
            model_text=model_text,
            brand_code_raw=effective_brand_code,
            brand_raw=brand_raw,
            category_code=resolved_category_code,
            models_by_code=models_by_code,
            models_by_code_unbranded=models_by_code_unbranded,
            models_by_name=models_by_name,
            ambiguous_model_codes=ambiguous_model_codes,
        )
        if reason:
            errors.append({"row": row_num, "reason": reason})
            continue
        item_name_norm = _normalize_item_name(item_name)
        week = _clean_value(_get(row, "周"))
        month = f"{year:04d}-{month_num:02d}"
        match_key_type = _match_key_type(item_id, item_url, item_name_norm)
        history_key = _history_identity_key(
            platform=platform,
            item_id=item_id,
            item_url=item_url,
            item_name_norm=item_name_norm,
            year=year,
            month_num=month_num,
            week=week,
        )

        values = {
            "import_batch": import_batch,
            "platform": platform,
            "item_id": item_id,
            "item_url": item_url,
            "item_name": item_name,
            "item_name_norm": item_name_norm,
            "year": year,
            "month_num": month_num,
            "week": week,
            "month": month,
            "report_type": _clean_value(_get(row, "报告类型")),
            "channel": _clean_value(_get(row, "渠道")),
            "category_name_raw": category_name_raw,
            "category_code_raw": category_code_raw,
            "brand_raw": brand_raw,
            "brand_code_raw": brand_code_raw or effective_brand_code,
            "model_text": model_text,
            "model_code_raw": model_code_raw,
            "model_id": model.id if model else None,
            "model_code": model.model_code if model else None,
            "category_code": model.category_code if model else resolved_category_code,
            "sales_amount": _clean_decimal(_get(row, "销额")),
            "sales_qty": _clean_int(_get(row, "销量")),
            "price": _clean_decimal(_get(row, "单价")),
            "match_key_type": match_key_type,
            "raw_payload": _json_safe_row(row),
            "updated_at": datetime.utcnow(),
        }

        if history_key in merged_rows:
            row_nums = merged_row_nums[history_key]
            row_nums.append(row_num)
            _merge_duplicate_history_values(merged_rows[history_key], values, row_nums)
        else:
            merged_rows[history_key] = values
            merged_row_nums[history_key] = [row_num]
        success += 1
        if len(merged_rows) >= 1000:
            flush_batch()

    flush_batch()
    return {"success": success, "created": created, "updated": updated, "errors": errors, "import_batch": import_batch}


def _import_historical_stream(
    db: Session,
    raw_rows_factory,
    mapping: dict[str, str],
    category_code: Optional[str],
    import_batch: str,
) -> dict:
    model_code_values, model_text_values = _model_lookup_values_from_rows(
        _standardize_historical_row(row, mapping, category_code) for row in raw_rows_factory()
    )
    models_by_code, models_by_code_unbranded, models_by_name = _load_preview_models(db, model_code_values, model_text_values)
    ambiguous_model_codes = _ambiguous_model_codes_from_rows(
        (_standardize_historical_row(row, mapping, category_code) for row in raw_rows_factory()),
        models_by_code_unbranded,
    )
    return _import_historical_rows(
        db,
        (_standardize_historical_row(row, mapping, category_code) for row in raw_rows_factory()),
        import_batch,
        models_by_code=models_by_code,
        models_by_code_unbranded=models_by_code_unbranded,
        models_by_name=models_by_name,
        ambiguous_model_codes=ambiguous_model_codes,
    )


@router.post("/import")
async def import_historical_mappings(file: UploadFile = File(...), db: Session = Depends(get_db)):
    _validate_excel_file(file)
    _, save_path, safe_filename = await save_tmp_file(file, settings.UPLOAD_DIR)
    try:
        xls = pd.ExcelFile(save_path)
        sheet_name, columns, mapping = _detect_sheet_and_mapping(xls)
        category_code = _infer_category_code(safe_filename, db)
        mapping = _normalize_mapping_for_columns(mapping, columns)
        usecols = _mapped_columns(mapping, columns)
        if save_path.suffix.lower() == ".xls":
            df = _read_sheet_preview(xls, sheet_name, usecols=usecols)
            standardized_df = _standardize_historical_df(df, mapping, category_code)
            return _import_historical_dataframe(db, standardized_df, safe_filename)
        return _import_historical_stream(
            db,
            lambda: _iter_sheet_rows_streaming(save_path, sheet_name, columns=columns, usecols=usecols),
            mapping,
            category_code,
            safe_filename,
        )
    finally:
        save_path.unlink(missing_ok=True)


@router.post("/headers")
async def historical_headers(file: UploadFile = File(...), db: Session = Depends(get_db)):
    _validate_excel_file(file)
    temp_file_id, save_path, safe_filename = await save_tmp_file(file, settings.UPLOAD_DIR)
    try:
        xls = pd.ExcelFile(save_path)
        sheet_name, columns, mapping = _detect_sheet_and_mapping(xls)
        category_code = _infer_category_code(safe_filename, db)
        return _historical_preview_response(
            temp_file_id=temp_file_id,
            filename=safe_filename,
            xls=xls,
            path=save_path,
            sheet_name=sheet_name,
            columns=columns,
            mapping=mapping,
            category_code=category_code,
            db=db,
        )
    except Exception:
        save_path.unlink(missing_ok=True)
        raise


@router.post("/preview")
def historical_preview(
    payload: HistoricalPreviewIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_historical_category_visible(db, current_user, payload.category_code)
    save_path = _find_tmp_file(payload.temp_file_id)
    xls = pd.ExcelFile(save_path)
    if payload.sheet_name not in xls.sheet_names:
        raise HTTPException(400, "选择的 sheet 不存在")
    df_header = _read_sheet_preview(xls, payload.sheet_name, nrows=0)
    columns = [str(col).strip() for col in df_header.columns]
    filename = _original_filename_from_tmp(save_path, payload.temp_file_id)
    return _historical_preview_response(
        temp_file_id=payload.temp_file_id,
        filename=filename,
        xls=xls,
        path=save_path,
        sheet_name=payload.sheet_name,
        columns=columns,
        mapping=payload.mapping,
        category_code=payload.category_code,
        db=db,
    )


@router.post("/confirm")
def historical_confirm(
    payload: HistoricalConfirmIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_historical_category_visible(db, current_user, payload.category_code)
    save_path = _find_tmp_file(payload.temp_file_id)
    xls = pd.ExcelFile(save_path)
    if payload.sheet_name not in xls.sheet_names:
        raise HTTPException(400, "选择的 sheet 不存在")
    df_header = _read_sheet_preview(xls, payload.sheet_name, nrows=0)
    columns = [str(col).strip() for col in df_header.columns]
    issues = _mapping_issues(payload.mapping, payload.category_code, columns)
    if issues:
        raise HTTPException(422, "；".join(issues))
    mapping = _normalize_mapping_for_columns(payload.mapping, columns)
    issues = _mapping_issues(mapping, payload.category_code, columns)
    if issues:
        raise HTTPException(422, "；".join(issues))
    usecols = _mapped_columns(mapping, columns)
    filename = _original_filename_from_tmp(save_path, payload.temp_file_id)
    if save_path.suffix.lower() == ".xls":
        df = _read_sheet_preview(xls, payload.sheet_name, usecols=usecols)
        standardized_df = _standardize_historical_df(df, mapping, payload.category_code)
        result = _import_historical_dataframe(db, standardized_df, filename)
    else:
        result = _import_historical_stream(
            db,
            lambda: _iter_sheet_rows_streaming(save_path, payload.sheet_name, columns=columns, usecols=usecols),
            mapping,
            payload.category_code,
            filename,
        )
    save_path.unlink(missing_ok=True)
    return result


@router.get("/batches")
def list_batches(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(
        HistoricalMapping.import_batch,
        func.count(HistoricalMapping.id).label("count"),
        func.max(HistoricalMapping.updated_at).label("updated_at"),
    ).filter(HistoricalMapping.import_batch.isnot(None))
    q = _filter_historical_visible_categories(q, db, current_user)
    rows = (
        q.group_by(HistoricalMapping.import_batch)
        .order_by(func.max(HistoricalMapping.updated_at).desc(), func.max(HistoricalMapping.id).desc())
        .all()
    )
    return [{"batch": r.import_batch, "count": r.count, "updated_at": format_beijing_datetime(r.updated_at)} for r in rows]


@router.get("/mappings")
def list_mappings(
    platform: Optional[str] = Query(None),
    import_batch: Optional[str] = Query(None),
    category_code: Optional[str] = Query(None),
    model_keyword: Optional[str] = Query(None),
    item_keyword: Optional[str] = Query(None),
    month: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = (
        db.query(HistoricalMapping, ModelRecord)
        .outerjoin(ModelRecord, HistoricalMapping.model_id == ModelRecord.id)
    )
    q = _filter_historical_visible_categories(q, db, current_user)
    if platform:
        q = q.filter(HistoricalMapping.platform == platform.lower())
    if import_batch:
        q = q.filter(HistoricalMapping.import_batch == import_batch)
    if category_code:
        q = q.filter(HistoricalMapping.category_code == category_code)
    if month:
        q = q.filter(HistoricalMapping.month == month)
    if model_keyword:
        pattern = f"%{model_keyword}%"
        q = q.filter(or_(
            HistoricalMapping.model_text.ilike(pattern),
            HistoricalMapping.model_code.ilike(pattern),
            ModelRecord.model_name.ilike(pattern),
        ))
    if item_keyword:
        pattern = f"%{item_keyword}%"
        q = q.filter(or_(
            HistoricalMapping.item_name.ilike(pattern),
            HistoricalMapping.item_id.ilike(pattern),
            HistoricalMapping.item_url.ilike(pattern),
        ))

    total = q.count()
    rows = (
        q.order_by(HistoricalMapping.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    items = [
        {
            "id": hm.id,
            "platform": hm.platform,
            "item_id": hm.item_id,
            "item_url": hm.item_url,
            "item_name": hm.item_name,
            "brand_raw": hm.brand_raw,
            "brand_code_raw": hm.brand_code_raw,
            "model_text": hm.model_text,
            "model_id": hm.model_id,
            "model_code": hm.model_code,
            "standard_model_name": m.model_name if m else None,
            "category_code": hm.category_code,
            "category_name_raw": hm.category_name_raw,
            "year": hm.year,
            "month_num": hm.month_num,
            "month": hm.month,
            "week": hm.week,
            "sales_qty": hm.sales_qty,
            "price": float(hm.price) if hm.price is not None else None,
            "sales_amount": float(hm.sales_amount) if hm.sales_amount is not None else None,
            "import_batch": hm.import_batch,
            "match_key_type": hm.match_key_type,
            "updated_at": format_beijing_datetime(hm.updated_at),
        }
        for hm, m in rows
    ]
    return {"total": total, "page": page, "page_size": page_size, "items": items}


@router.delete("/mappings/batch", status_code=204)
def delete_batch(
    body: BatchDeleteIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(HistoricalMapping).filter(HistoricalMapping.import_batch == body.import_batch)
    q = _filter_historical_visible_categories(q, db, current_user)
    deleted = q.delete(synchronize_session=False)
    if deleted == 0:
        raise HTTPException(404, f"批次 '{body.import_batch}' 不存在或已删除")
    db.commit()


@router.delete("/mappings/{mapping_id}", status_code=204)
def delete_mapping(
    mapping_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(HistoricalMapping).filter(HistoricalMapping.id == mapping_id)
    q = _filter_historical_visible_categories(q, db, current_user)
    row = q.first()
    if not row:
        raise HTTPException(404, "记录不存在")
    db.delete(row)
    db.commit()
