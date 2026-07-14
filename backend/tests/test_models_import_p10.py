"""Tests for models import P10 headers/confirm endpoints."""
import io
import os
import tempfile
import openpyxl
import pytest
from pathlib import Path
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from fastapi import FastAPI
from fastapi.testclient import TestClient

from app.models.database import Base, get_db
from app.models.schemas import Category, ModelRecord, ModelSpec
from app.api.models_api import router


# ─── Fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture()
def tmp_upload(tmp_path):
    """Create a temp upload directory and patch UPLOAD_DIR in models_api."""
    import app.api.models_api as models_mod
    original = models_mod.UPLOAD_DIR
    models_mod.UPLOAD_DIR = str(tmp_path)
    yield tmp_path
    models_mod.UPLOAD_DIR = original


@pytest.fixture()
def client(tmp_upload):
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)

    # Pre-seed a category
    s = Session()
    s.add(Category(code="CAT001", name="测试品类"))
    s.commit()
    s.close()

    def override_db():
        db = Session()
        try:
            yield db
        finally:
            db.close()

    app = FastAPI()
    app.include_router(router)
    app.dependency_overrides[get_db] = override_db
    return TestClient(app)


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _make_xlsx(headers, rows=None, sheet_name="Sheet") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for r in (rows or []):
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_headers(client, headers, rows=None):
    """POST to /api/models/headers and return response."""
    xlsx_bytes = _make_xlsx(headers, rows)
    return client.post(
        "/api/models/headers",
        files={"file": ("test.xlsx", xlsx_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def _make_models_template_xlsx(model_rows, spec_rows) -> bytes:
    wb = openpyxl.Workbook()
    model_ws = wb.active
    model_ws.title = "型号"
    model_ws.append(["品牌码", "型号码", "品类", "品牌名称", "型号名称", "上市年", "上市月", "上市周", "上市价格", "网址"])
    for row in model_rows:
        model_ws.append(row)

    spec_ws = wb.create_sheet("型号规格")
    spec_ws.append(["品牌码", "型号码", "规格名称", "规格值"])
    for row in spec_rows:
        spec_ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _headers_then_confirm(client, headers, data_rows, mapping, category_code="CAT001", extra_payload=None):
    """Two-step: upload headers, then confirm."""
    resp = _upload_headers(client, headers, data_rows)
    assert resp.status_code == 200
    temp_file_id = resp.json()["temp_file_id"]

    payload = {
        "temp_file_id": temp_file_id,
        "mapping": mapping,
        "ignore_columns": [],
        "category_code": category_code,
    }
    if extra_payload:
        payload.update(extra_payload)

    return client.post("/api/models/confirm", json=payload)


# ─── headers ──────────────────────────────────────────────────────────────────

def test_models_headers_returns_columns(client):
    resp = _upload_headers(client, ["brand_code", "model_code", "brand_name"])
    assert resp.status_code == 200
    data = resp.json()
    assert "temp_file_id" in data
    assert "brand_code" in data["columns"]
    assert "model_code" in data["columns"]


def test_models_headers_requires_file(client):
    resp = client.post("/api/models/headers")
    assert resp.status_code == 422


def test_models_headers_returns_filename(client):
    resp = _upload_headers(client, ["brand_code", "model_code"])
    assert resp.status_code == 200
    data = resp.json()
    assert "filename" in data
    assert data["filename"] == "test.xlsx"


def test_models_headers_suggests_template_when_match(client):
    """Headers endpoint suggests existing template when columns match."""
    from app.services.import_helper import col_fingerprint
    from app.models.schemas import ColumnTemplate

    # Insert matching template into DB
    db_gen = client.app.dependency_overrides[get_db]()
    db = next(db_gen)
    cols = ["brand_code", "model_code"]
    fp = col_fingerprint(cols)
    tmpl = ColumnTemplate(
        name="型号模板",
        module="model",
        mapping={"brand_code": "brand_code", "model_code": "model_code"},
        ignore_columns=[],
        col_fingerprint=fp,
    )
    db.add(tmpl)
    db.commit()
    db.close()

    resp = _upload_headers(client, cols)
    assert resp.status_code == 200
    data = resp.json()
    assert data.get("suggested_template") is not None
    assert data["match_score"] == 100


# ─── confirm ──────────────────────────────────────────────────────────────────

def test_models_confirm_inserts_model(client):
    resp = _headers_then_confirm(
        client,
        headers=["brand_code", "model_code"],
        data_rows=[["BRAND_A", "MODEL_X"]],
        mapping={"brand_code": "brand_code", "model_code": "model_code"},
        category_code="CAT001",
    )
    assert resp.status_code == 200
    d = resp.json()
    assert "models_inserted" in d
    assert "errors" in d
    assert d["models_inserted"] == 1


def test_models_confirm_category_fallback(client):
    """category_code from payload is used when Excel row has no category_code col."""
    resp = _headers_then_confirm(
        client,
        headers=["brand_code", "model_code"],
        data_rows=[["BR1", "M1"]],
        mapping={"brand_code": "brand_code", "model_code": "model_code"},
        category_code="FALLBACK_CAT",
    )
    assert resp.status_code == 200
    d = resp.json()
    assert d["models_inserted"] == 1


def test_models_confirm_selected_category_overrides_excel_category(client):
    """Import should use the category selected in step 1, not a stale Excel category value."""
    resp = _headers_then_confirm(
        client,
        headers=["brand_code", "model_code", "category_code"],
        data_rows=[["BR1", "M1", "STALE_CAT"]],
        mapping={
            "brand_code": "brand_code",
            "model_code": "model_code",
            "category_code": "category_code",
        },
        category_code="FALLBACK_CAT",
    )
    assert resp.status_code == 200
    db = next(client.app.dependency_overrides[get_db]())
    try:
        model = db.query(ModelRecord).filter_by(brand_code="BR1", model_code="M1").one()
        assert model.category_code == "FALLBACK_CAT"
    finally:
        db.close()


def test_models_confirm_missing_required_field(client):
    """Rows missing brand_code or model_code are counted as errors."""
    resp = _headers_then_confirm(
        client,
        headers=["brand_code"],
        data_rows=[["BRAND_A"]],
        mapping={"brand_code": "brand_code"},
        category_code="CAT001",
    )
    assert resp.status_code == 200
    d = resp.json()
    assert d["errors"] or d["models_inserted"] == 0


def test_models_confirm_upserts_on_duplicate(client):
    """Importing the same brand_code+model_code updates (models_updated) instead of inserting."""
    mapping = {"brand_code": "brand_code", "model_code": "model_code"}
    headers = ["brand_code", "model_code"]
    rows = [["BRAND_A", "MODEL_Y"]]

    resp1 = _headers_then_confirm(client, headers, rows, mapping)
    assert resp1.status_code == 200
    assert resp1.json()["models_inserted"] == 1

    resp2 = _headers_then_confirm(client, headers, rows, mapping)
    assert resp2.status_code == 200
    d2 = resp2.json()
    assert d2["models_updated"] == 1
    assert d2["models_inserted"] == 0


def test_models_headers_suggests_builtin_template_mapping_for_downloaded_template(client):
    mapping = {
        "品牌码": "brand_code",
        "型号码": "model_code",
        "品类": "category_code",
        "品牌名称": "brand_name",
        "型号名称": "model_name",
        "上市年": "launch_year",
        "上市月": "launch_month",
        "上市周": "launch_week",
        "上市价格": "launch_price",
        "网址": "url",
    }

    xlsx_bytes = _make_models_template_xlsx(
        model_rows=[
            ["DJI", "OSMO-ACTION-4", "CAT001", "大疆", "Osmo Action 4", 2024, 9, None, 2999, "https://example.com/product"],
        ],
        spec_rows=[
            ["DJI", "OSMO-ACTION-4", "产品形态", "OA传统"],
        ],
    )
    resp = client.post(
        "/api/models/headers",
        files={"file": ("models.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert resp.status_code == 200
    data = resp.json()
    assert data["match_score"] == 100
    assert data["suggested_template"]["name"] == "产品属性导入模板"
    assert data["suggested_template"]["mapping"] == mapping


def test_models_confirm_imports_specs_from_template_second_sheet(client):
    xlsx_bytes = _make_models_template_xlsx(
        model_rows=[
            ["DJI", "OSMO-ACTION-4", "CAT001", "大疆", "Osmo Action 4", 2024, 9, None, 2999, "https://example.com/product"],
        ],
        spec_rows=[
            ["DJI", "OSMO-ACTION-4", "产品形态", "OA传统"],
            ["不需要填写", "不需要填写", "防抖", "电子防抖"],
        ],
    )
    headers_resp = client.post(
        "/api/models/headers",
        files={"file": ("models.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert headers_resp.status_code == 200

    confirm_resp = client.post(
        "/api/models/confirm",
        json={
            "temp_file_id": headers_resp.json()["temp_file_id"],
            "mapping": {
                "品牌码": "brand_code",
                "型号码": "model_code",
                "品类": "category_code",
                "品牌名称": "brand_name",
                "型号名称": "model_name",
                "上市年": "launch_year",
                "上市月": "launch_month",
                "上市周": "launch_week",
                "上市价格": "launch_price",
                "网址": "url",
            },
            "ignore_columns": [],
            "category_code": "CAT001",
        },
    )

    assert confirm_resp.status_code == 200
    data = confirm_resp.json()
    assert data["models_inserted"] == 1
    assert data["specs_inserted"] == 2

    db = next(client.app.dependency_overrides[get_db]())
    try:
        model = db.query(ModelRecord).filter_by(brand_code="DJI", model_code="OSMO-ACTION-4").one()
        specs = db.query(ModelSpec).filter_by(model_id=model.id).order_by(ModelSpec.spec_name).all()
        assert [(s.spec_name, s.spec_value) for s in specs] == [("产品形态", "OA传统"), ("防抖", "电子防抖")]
    finally:
        db.close()


def test_models_confirm_replaces_existing_specs_for_uploaded_model(client):
    db = next(client.app.dependency_overrides[get_db]())
    try:
        model = ModelRecord(
            brand_code="DJI",
            model_code="OSMO-ACTION-4",
            category_code="CAT001",
            brand_name="大疆",
            model_name="Old Osmo Action 4",
        )
        db.add(model)
        db.flush()
        db.add_all([
            ModelSpec(model_id=model.id, spec_name="旧规格", spec_value="旧值"),
            ModelSpec(model_id=model.id, spec_name="防抖", spec_value="旧防抖"),
        ])
        db.commit()
    finally:
        db.close()

    xlsx_bytes = _make_models_template_xlsx(
        model_rows=[
            ["DJI", "OSMO-ACTION-4", "CAT001", "大疆", "Osmo Action 4", 2024, 9, None, 2999, "https://example.com/product"],
        ],
        spec_rows=[
            ["DJI", "OSMO-ACTION-4", "产品形态", "OA传统"],
            ["不需要填写", "不需要填写", "防抖", "电子防抖"],
        ],
    )
    headers_resp = client.post(
        "/api/models/headers",
        files={"file": ("models.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert headers_resp.status_code == 200

    confirm_resp = client.post(
        "/api/models/confirm",
        json={
            "temp_file_id": headers_resp.json()["temp_file_id"],
            "mapping": {
                "品牌码": "brand_code",
                "型号码": "model_code",
                "品类": "category_code",
                "品牌名称": "brand_name",
                "型号名称": "model_name",
                "上市年": "launch_year",
                "上市月": "launch_month",
                "上市周": "launch_week",
                "上市价格": "launch_price",
                "网址": "url",
            },
            "ignore_columns": [],
            "category_code": "CAT001",
        },
    )

    assert confirm_resp.status_code == 200
    data = confirm_resp.json()
    assert data["models_updated"] == 1
    assert data["specs_inserted"] == 2

    db = next(client.app.dependency_overrides[get_db]())
    try:
        model = db.query(ModelRecord).filter_by(brand_code="DJI", model_code="OSMO-ACTION-4").one()
        specs = db.query(ModelSpec).filter_by(model_id=model.id).order_by(ModelSpec.spec_name).all()
        spec_pairs = [(s.spec_name, s.spec_value) for s in specs]
        assert spec_pairs == [("产品形态", "OA传统"), ("防抖", "电子防抖")]
        assert ("旧规格", "旧值") not in spec_pairs
        assert ("防抖", "旧防抖") not in spec_pairs
        assert len(spec_pairs) == 2
    finally:
        db.close()



def test_models_confirm_temp_file_not_found(client):
    """Confirm returns 404 when temp_file_id does not exist."""
    payload = {
        "temp_file_id": "nonexistent-uuid",
        "mapping": {"brand_code": "brand_code", "model_code": "model_code"},
        "ignore_columns": [],
        "category_code": "CAT001",
    }
    resp = client.post("/api/models/confirm", json=payload)
    assert resp.status_code == 404


def test_models_confirm_multiple_rows(client):
    """Confirm inserts multiple valid rows."""
    resp = _headers_then_confirm(
        client,
        headers=["brand_code", "model_code"],
        data_rows=[
            ["BRAND_A", "MODEL_1"],
            ["BRAND_A", "MODEL_2"],
            ["BRAND_B", "MODEL_3"],
        ],
        mapping={"brand_code": "brand_code", "model_code": "model_code"},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["models_inserted"] == 3
    assert data["errors"] == []


def test_models_confirm_column_remapping(client):
    """Confirm supports remapping non-standard column names."""
    resp = _headers_then_confirm(
        client,
        headers=["品牌码", "型号码"],
        data_rows=[["SONY", "WH1000XM5"]],
        mapping={"品牌码": "brand_code", "型号码": "model_code"},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["models_inserted"] == 1
