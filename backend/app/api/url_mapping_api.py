# backend/app/api/url_mapping_api.py
"""URL→型号映射表 CRUD + Excel批量导入"""
from datetime import datetime
from typing import Optional
import io
import os
from pathlib import Path

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from pydantic import BaseModel
from sqlalchemy import or_
from sqlalchemy.orm import Session, joinedload
import openpyxl

from app.models.database import get_db
from app.models.schemas import (
    ItemUrlMapping, ItemUrlMappingIn, ItemUrlMappingOut,
    ModelRecord, PaginatedResponse,
)
from app.utils.url_utils import extract_item_id
from app.services.import_helper import save_tmp_file, read_columns, find_best_template, col_fingerprint

router = APIRouter(prefix="/api/url-mappings", tags=["url-mappings"])

UPLOAD_DIR = os.environ.get("UPLOAD_DIR", "./uploads")

# 渠道名称 → platform 规范值
_PLATFORM_MAP = {
    "JD": "jd", "京东": "jd",
    "TMALL": "tmall", "天猫": "tmall",
    "TAOBAO": "taobao", "淘宝": "taobao",
    "SUNING": "suning", "苏宁": "suning",
}


def _to_out(m: ItemUrlMapping) -> ItemUrlMappingOut:
    model = m.model
    category = getattr(model, "category", None) if model else None
    return ItemUrlMappingOut(
        id=m.id,
        platform=m.platform,
        item_id=m.item_id,
        item_url=m.item_url,
        model_id=m.model_id,
        price=float(m.price) if m.price is not None else None,
        brand_code=model.brand_code if model else m.brand_code,
        model_code=model.model_code if model else None,
        brand_name=model.brand_name if model else None,
        model_name=model.model_name if model else None,
        category_code=model.category_code if model else None,
        category_name=category.name if category else None,
        item_name=None,
        source=m.source,
        data_year=m.data_year,
        data_month=m.data_month,
        operator=None,
        created_at=m.created_at,
        updated_at=m.updated_at,
    )


class UrlMappingConfirmPayload(BaseModel):
    temp_file_id: str
    mapping: dict
    ignore_columns: list = []
    category_code: str
    save_template_name: Optional[str] = None
    data_year: Optional[int] = None
    data_month: Optional[int] = None


@router.post("/headers")
async def url_mapping_headers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """P10: Step 1 — read columns, suggest mapping template."""
    temp_file_id, save_path, filename = await save_tmp_file(file, UPLOAD_DIR)
    columns = read_columns(save_path)
    best_tmpl, score = find_best_template(columns, "url", db)
    return {
        "temp_file_id": temp_file_id,
        "filename": filename,
        "columns": columns,
        "suggested_template": {
            "id": best_tmpl.id,
            "name": best_tmpl.name,
            "mapping": best_tmpl.mapping,
            "ignore_columns": best_tmpl.ignore_columns or [],
        } if best_tmpl else None,
        "match_score": score,
    }


@router.post("/confirm")
def url_mapping_confirm(
    payload: UrlMappingConfirmPayload,
    db: Session = Depends(get_db),
):
    """P10: Step 2 — parse file with mapping, upsert item_url_mappings."""
    from app.models.schemas import ItemUrlMapping, ModelRecord as ModelORM

    tmp_dir = Path(UPLOAD_DIR) / "tmp"
    candidates = list(tmp_dir.glob(f"*{payload.temp_file_id}*")) if tmp_dir.exists() else []
    if not candidates:
        direct = tmp_dir / payload.temp_file_id
        if direct.exists():
            candidates = [direct]
    if not candidates:
        raise HTTPException(status_code=404, detail="Temp file not found or expired")
    file_path = candidates[0]

    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        try:
            df = pd.read_csv(file_path, dtype=str, encoding="utf-8-sig")
        except Exception:
            df = pd.read_csv(file_path, dtype=str, encoding="gbk")
    else:
        df = pd.read_excel(file_path, sheet_name=0, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    ignore_set = set(payload.ignore_columns)
    rename_map = {
        col: target
        for col, target in payload.mapping.items()
        if col in df.columns and col not in ignore_set
    }
    df = df.rename(columns=rename_map)

    inserted = 0
    updated = 0
    skipped = 0
    errors = []

    for i, row in enumerate(df.itertuples(index=False), start=2):
        row_dict = row._asdict()

        platform_raw = str(row_dict.get("platform") or "").strip()
        item_url = str(row_dict.get("item_url") or "").strip()
        brand_code = str(row_dict.get("brand_code") or "").strip()
        model_code = str(row_dict.get("model_code") or "").strip()

        if not platform_raw or not item_url or not brand_code or not model_code:
            errors.append(f"Row {i}: missing required field")
            continue

        # Normalize platform: try map lookup (handles Chinese/uppercase), fallback to lowercase
        platform = _PLATFORM_MAP.get(platform_raw.upper(), platform_raw.lower())

        # Extract item_id from URL
        url_info = extract_item_id(item_url)
        if not url_info:
            errors.append(f"Row {i}: cannot extract item_id from URL '{item_url}'")
            continue
        url_platform, item_id = url_info

        # Lookup model
        model = (
            db.query(ModelORM)
            .filter(ModelORM.brand_code == brand_code, ModelORM.model_code == model_code)
            .first()
        )
        if not model:
            errors.append(f"Row {i}: model ({brand_code}, {model_code}) not found")
            continue

        # Category mismatch warning (non-blocking)
        if model.category_code and payload.category_code and model.category_code != payload.category_code:
            errors.append(
                f"Row {i}: warning — model category ({model.category_code}) "
                f"differs from selected ({payload.category_code})"
            )

        price_raw = row_dict.get("price")
        try:
            price = float(str(price_raw).strip()) if price_raw and str(price_raw).strip() not in ("", "nan", "None") else None
        except (ValueError, TypeError):
            price = None

        existing = (
            db.query(ItemUrlMapping)
            .filter(ItemUrlMapping.platform == url_platform, ItemUrlMapping.item_id == item_id)
            .first()
        )
        if existing:
            existing.model_id = model.id
            if price is not None:
                existing.price = price
            existing.source = 'url_import'
            existing.data_year = payload.data_year
            existing.data_month = payload.data_month
            updated += 1
        else:
            db.add(ItemUrlMapping(
                platform=url_platform,
                item_id=item_id,
                item_url=item_url,
                model_id=model.id,
                price=price,
                source='url_import',
                data_year=payload.data_year,
                data_month=payload.data_month,
            ))
            inserted += 1

    if inserted > 0 or updated > 0:
        db.commit()

    if payload.save_template_name and (inserted + updated) > 0:
        from app.models.schemas import ColumnTemplate
        tmpl = ColumnTemplate(
            name=payload.save_template_name,
            module="url",
            mapping=payload.mapping,
            ignore_columns=payload.ignore_columns,
            col_fingerprint=col_fingerprint(list(payload.mapping.keys())),
        )
        db.add(tmpl)
        try:
            db.commit()
        except Exception:
            db.rollback()

    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


@router.post("/import")
def import_url_mappings(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    从 Excel rawdata sheet 批量导入 URL→型号 映射。
    期望列：渠道 / 网址 / 品牌码 / 型号码 / 单价
    采用 upsert（platform+item_id 冲突时更新 model_id 和 price）。
    """
    contents = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)

    # 优先找名为 rawdata 的 sheet，否则取第一个
    sheet = wb["rawdata"] if "rawdata" in wb.sheetnames else wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "skipped": 0, "errors": []}

    # 构建列名→列索引映射（第一行为表头）
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col = {name: idx for idx, name in enumerate(header)}

    # 必需列：品牌码/型号码 可以用 品牌/型号 替代（rawdata sheet 有时只有后者）
    required = {"渠道", "网址"}
    required_brand = {"品牌码", "品牌"}   # 至少有其中一列
    required_model = {"型号码", "型号"}   # 至少有其中一列
    missing = required - set(col.keys())
    if missing:
        raise HTTPException(400, detail=f"Excel 缺少必需列：{missing}")
    if not (required_brand & set(col.keys())):
        raise HTTPException(400, detail="Excel 缺少品牌列（需要「品牌码」或「品牌」）")
    if not (required_model & set(col.keys())):
        raise HTTPException(400, detail="Excel 缺少型号列（需要「型号码」或「型号」）")

    # 构建 (brand_code, model_code) → model_id 缓存
    all_models = db.query(ModelRecord).all()
    model_lookup: dict[tuple[str, str], int] = {
        (m.brand_code.upper().strip(), m.model_code.upper().strip()): m.id
        for m in all_models
    }

    imported = 0
    skipped = 0
    errors: list[str] = []
    seen_keys: set[tuple[str, str]] = set()  # 同一文件内去重，避免重复 INSERT

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            platform_raw = str(row[col["渠道"]] or "").strip().upper()
            platform = _PLATFORM_MAP.get(platform_raw)
            url = str(row[col["网址"]] or "").strip()
            brand_code_raw = row[col["品牌码"]] if "品牌码" in col else None
            if not brand_code_raw and "品牌" in col:
                brand_code_raw = row[col["品牌"]]
            brand_code = str(brand_code_raw or "").strip().upper()

            model_code_raw = row[col["型号码"]] if "型号码" in col else None
            if not model_code_raw and "型号" in col:
                model_code_raw = row[col["型号"]]
            model_code = str(model_code_raw or "").strip().upper()
            price_raw = row[col["单价"]] if "单价" in col else None
            price = float(price_raw) if price_raw not in (None, "") else None
        except Exception as e:
            errors.append(f"第{row_idx}行解析错误：{e}")
            skipped += 1
            continue

        if not platform or not url or not brand_code or not model_code:
            skipped += 1
            continue

        url_info = extract_item_id(url)
        if not url_info or url_info[0] != platform:
            skipped += 1
            continue

        _, item_id = url_info
        key = (platform, item_id)

        # 同文件内重复行：只保留最后一条（继续处理，覆盖 seen_keys）
        if key in seen_keys:
            skipped += 1
            continue
        seen_keys.add(key)

        model_id = model_lookup.get((brand_code, model_code))
        if not model_id:
            errors.append(f"第{row_idx}行：型号 [{brand_code}]{model_code} 不存在，已跳过")
            skipped += 1
            continue

        # Upsert
        existing = db.query(ItemUrlMapping).filter_by(
            platform=platform, item_id=item_id
        ).first()
        if existing:
            existing.model_id = model_id
            existing.price = price
            existing.item_url = url
            existing.updated_at = datetime.utcnow()
            existing.source = 'url_import'
        else:
            db.add(ItemUrlMapping(
                platform=platform, item_id=item_id, item_url=url, model_id=model_id, price=price,
                source='url_import',
            ))
        imported += 1

    db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors[:20]}


@router.get("", response_model=PaginatedResponse)
def list_url_mappings(
    keyword: Optional[str] = Query(None),
    platform: Optional[str] = Query(None),
    category_code: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
):
    q = db.query(ItemUrlMapping).filter(ItemUrlMapping.brand_code.isnot(None))
    if platform:
        q = q.filter(ItemUrlMapping.platform == platform)
    if year is not None:
        q = q.filter(ItemUrlMapping.data_year == year)
    if month is not None:
        q = q.filter(ItemUrlMapping.data_month == month)
    if keyword:
        kw = f"%{keyword}%"
        q = q.outerjoin(ModelRecord, ItemUrlMapping.model_id == ModelRecord.id).filter(
            or_(
                ItemUrlMapping.item_id.ilike(kw),
                ModelRecord.model_code.ilike(kw),
                ModelRecord.brand_code.ilike(kw),
            )
        )
    elif category_code:
        q = q.outerjoin(ModelRecord, ItemUrlMapping.model_id == ModelRecord.id)
    if category_code:
        q = q.filter(ModelRecord.category_code == category_code)
    total = q.count()
    rows = (
        q.options(joinedload(ItemUrlMapping.model).joinedload(ModelRecord.category))
        .order_by(ItemUrlMapping.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return PaginatedResponse(
        total=total, page=page, page_size=page_size,
        items=[_to_out(r) for r in rows],
    )


@router.post("", response_model=ItemUrlMappingOut)
def create_url_mapping(payload: ItemUrlMappingIn, db: Session = Depends(get_db)):
    model_for_brand = db.query(ModelRecord).filter_by(id=payload.model_id).first() if payload.model_id else None
    if not model_for_brand:
        raise HTTPException(404, "型号不存在")
    existing = db.query(ItemUrlMapping).filter_by(
        platform=payload.platform, item_id=payload.item_id
    ).first()
    if existing:
        raise HTTPException(409, "该 platform+item_id 已存在，请使用编辑功能")
    m = ItemUrlMapping(
        platform=payload.platform,
        item_id=payload.item_id,
        item_url=payload.item_url,
        brand_code=model_for_brand.brand_code if model_for_brand else payload.brand_code,
        model_id=payload.model_id,
        price=payload.price,
        source='manual',
    )
    db.add(m)
    db.commit()
    db.refresh(m)
    return _to_out(m)


@router.put("/{mapping_id}", response_model=ItemUrlMappingOut)
def update_url_mapping(mapping_id: int, payload: ItemUrlMappingIn, db: Session = Depends(get_db)):
    m = db.query(ItemUrlMapping).filter_by(id=mapping_id).first()
    if not m:
        raise HTTPException(404, "映射记录不存在")
    model_for_brand = db.query(ModelRecord).filter_by(id=payload.model_id).first() if payload.model_id else None
    if not model_for_brand:
        raise HTTPException(404, "型号不存在")
    m.platform = payload.platform
    m.item_id = payload.item_id
    m.item_url = payload.item_url
    m.brand_code = model_for_brand.brand_code if model_for_brand else payload.brand_code
    m.model_id = payload.model_id
    m.price = payload.price
    m.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(m)
    return _to_out(m)


@router.delete("/{mapping_id}")
def delete_url_mapping(mapping_id: int, db: Session = Depends(get_db)):
    m = db.query(ItemUrlMapping).filter_by(id=mapping_id).first()
    if not m:
        raise HTTPException(404, "映射记录不存在")
    db.delete(m)
    db.commit()
    return {"deleted": True}
