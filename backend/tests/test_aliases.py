"""
model_aliases API 测试（SQLite 内存库）
"""
import pytest
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from app.models.database import Base
from app.models.schemas import ModelRecord, ModelAlias


@pytest.fixture
def db():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )

    @event.listens_for(engine, "connect")
    def set_pragma(conn, _):
        conn.execute("PRAGMA foreign_keys=ON")

    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()


def test_model_alias_orm_exists(db):
    """ModelAlias ORM 可以写入和查询。"""
    m = ModelRecord(brand_code="SONY", model_code="HT-A7000",
                    brand_name="索尼")
    db.add(m)
    db.flush()

    alias = ModelAlias(model_id=m.id, alias_code="HTA7000")
    db.add(alias)
    db.commit()

    result = db.query(ModelAlias).filter(ModelAlias.model_id == m.id).all()
    assert len(result) == 1
    assert result[0].alias_code == "HTA7000"


from fastapi import FastAPI
from fastapi.testclient import TestClient
from app.api.models_api import router
from app.models.database import get_db


def _make_client(session):
    app = FastAPI()
    app.include_router(router)

    def _get():
        yield session

    app.dependency_overrides[get_db] = _get
    return TestClient(app)


def test_add_and_list_aliases(db):
    """POST /api/models/{id}/aliases 新增别名，GET 返回列表。"""
    client = _make_client(db)

    m = ModelRecord(brand_code="SONY", model_code="HT-X9000F",
                    brand_name="索尼")
    db.add(m)
    db.commit()

    res = client.post(f"/api/models/{m.id}/aliases", json={"alias_code": "HTX9000F"})
    assert res.status_code == 200
    data = res.json()
    assert data["alias_code"] == "HTX9000F"
    assert "id" in data

    res2 = client.get(f"/api/models/{m.id}/aliases")
    assert res2.status_code == 200
    items = res2.json()
    assert len(items) == 1
    assert items[0]["alias_code"] == "HTX9000F"


def test_delete_alias(db):
    """DELETE /api/models/{id}/aliases/{alias_id} 删除别名。"""
    client = _make_client(db)

    m = ModelRecord(brand_code="SONY", model_code="HT-S400",
                    brand_name="索尼")
    db.add(m)
    db.flush()
    alias = ModelAlias(model_id=m.id, alias_code="HTS400")
    db.add(alias)
    db.commit()

    res = client.delete(f"/api/models/{m.id}/aliases/{alias.id}")
    assert res.status_code == 200

    remaining = db.query(ModelAlias).filter(ModelAlias.model_id == m.id).all()
    assert len(remaining) == 0


import io
import openpyxl


def _make_excel_with_alias() -> bytes:
    """构造含「别名」sheet 的 Excel，返回字节内容。"""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "型号"
    ws1.append(["品牌码", "型号码"])
    ws1.append(["SONY", "HT-A3000"])
    wb.create_sheet("型号规格")
    ws3 = wb.create_sheet("别名")
    ws3.append(["品牌码", "型号码", "别名"])
    ws3.append(["SONY", "HT-A3000", "HTA3000"])
    ws3.append(["SONY", "HT-A3000", "HT A3000"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_import_aliases_from_excel(db):
    """Excel「别名」sheet 中的别名在导入后写入 model_aliases 表。"""
    client = _make_client(db)

    content = _make_excel_with_alias()
    res = client.post(
        "/api/models/import",
        files={"file": ("models.xlsx", content,
               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    data = res.json()
    assert data["imported_models"] == 1
    assert data["imported_aliases"] == 2

    m = db.query(ModelRecord).filter(ModelRecord.model_code == "HT-A3000").first()
    assert m is not None
    aliases = db.query(ModelAlias).filter(ModelAlias.model_id == m.id).all()
    alias_codes = {a.alias_code for a in aliases}
    assert "HTA3000" in alias_codes
    assert "HT A3000" in alias_codes
