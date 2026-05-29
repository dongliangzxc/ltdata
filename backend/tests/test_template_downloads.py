from contextlib import asynccontextmanager
import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook


@asynccontextmanager
async def _noop_lifespan(application):
    yield


def _client(monkeypatch):
    from app.main import app

    monkeypatch.setattr("app.core.security.verify_token", lambda token: "test_user")
    monkeypatch.setattr("app.main.verify_token", lambda token: "test_user")
    monkeypatch.setattr(app.router, "lifespan_context", _noop_lifespan)
    return TestClient(app, raise_server_exceptions=True)


def test_metadata_template_download_uses_backend_packaged_file(monkeypatch):
    from app.api.metadata import _METADATA_TEMPLATE_PATH

    assert _METADATA_TEMPLATE_PATH.name == "洛图科技—产品段属性说明-模板.xlsx"
    assert _METADATA_TEMPLATE_PATH.parent.name == "templates"
    assert _METADATA_TEMPLATE_PATH.parent.parent.name == "app"


def test_metadata_template_download_returns_existing_excel_file(monkeypatch):
    client = _client(monkeypatch)
    resp = client.get("/api/metadata/template", headers={"Authorization": "Bearer test_token"})

    assert resp.status_code == 200
    assert resp.content
    assert "spreadsheetml.sheet" in resp.headers["content-type"]
    assert "filename" in resp.headers["content-disposition"]


def test_metadata_template_can_be_previewed_for_import(monkeypatch):
    client = _client(monkeypatch)
    template_resp = client.get("/api/metadata/template", headers={"Authorization": "Bearer test_token"})

    resp = client.post(
        "/api/metadata/preview",
        files={
            "file": (
                "洛图科技—产品段属性说明-模板.xlsx",
                template_resp.content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers={"Authorization": "Bearer test_token"},
    )

    assert resp.status_code == 200
    data = resp.json()
    assert data["total_rows"] > 0
    assert data["valid_rows"] > 0
    assert data["preview"][0]["spec_name"]


def test_models_template_download_returns_two_sheet_workbook(monkeypatch):
    client = _client(monkeypatch)
    resp = client.get("/api/models/template", headers={"Authorization": "Bearer test_token"})

    assert resp.status_code == 200
    assert resp.content

    workbook = load_workbook(io.BytesIO(resp.content))
    assert workbook.sheetnames == ["型号", "型号规格"]
    assert [cell.value for cell in workbook["型号"][1]] == [
        "品牌码",
        "型号码",
        "品类",
        "品牌名称",
        "型号名称",
        "上市年",
        "上市月",
        "上市周",
        "上市价格",
        "网址",
    ]
    assert [cell.value for cell in workbook["型号规格"][1]] == [
        "品牌码",
        "型号码",
        "规格名称",
        "规格值",
    ]
